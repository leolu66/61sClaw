import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv('billing_2026-02-01_2026-02-26.csv')
df.columns = ['date', 'model', 'requests', 'input_tokens', 'output_tokens', 'total_tokens', 'billing_type', 'price_detail', 'cost']
df['date'] = pd.to_datetime(df['date'])

wb = Workbook()

# Sheet1: Raw Data
ws1 = wb.active
ws1.title = 'RawData'
for r in dataframe_to_rows(df, index=False, header=True):
    ws1.append(r)

# Sheet2: By Model
ws2 = wb.create_sheet('ByModel')
model_summary = df.groupby('model').agg({
    'requests': 'sum',
    'input_tokens': 'sum',
    'output_tokens': 'sum',
    'total_tokens': 'sum',
    'cost': 'sum'
}).sort_values('cost', ascending=False).reset_index()

ws2.append(['Model', 'Requests', 'InputTokens', 'OutputTokens', 'TotalTokens', 'Cost(CNY)'])
for _, row in model_summary.iterrows():
    ws2.append([row['model'], row['requests'], row['input_tokens'], row['output_tokens'], row['total_tokens'], round(row['cost'], 2)])

ws2['H1'] = 'Model'
ws2['I1'] = 'Pct'
model_summary['pct'] = model_summary['cost'] / model_summary['cost'].sum() * 100
for i, row in model_summary.iterrows():
    ws2[f'H{i+2}'] = row['model']
    ws2[f'I{i+2}'] = round(row['pct'], 1)

pie = PieChart()
labels = Reference(ws2, min_col=8, min_row=2, max_row=len(model_summary)+1)
data = Reference(ws2, min_col=9, min_row=1, max_row=len(model_summary)+1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = 'Cost by Model'
pie.height = 10
pie.width = 15
ws2.add_chart(pie, 'K1')

# Sheet2b: By Model - All 3 Metrics
ws2b = wb.create_sheet('ByModel_AllMetrics')
ws2b.append(['Model', 'Requests', 'TotalTokens', 'Cost(CNY)', 'RequestsPct', 'TokensPct', 'CostPct'])
total_reqs = model_summary['requests'].sum()
total_toks = model_summary['total_tokens'].sum()
total_cost = model_summary['cost'].sum()
for _, row in model_summary.iterrows():
    req_pct = round(row['requests'] / total_reqs * 100, 1)
    tok_pct = round(row['total_tokens'] / total_toks * 100, 1)
    cost_pct = round(row['cost'] / total_cost * 100, 1)
    ws2b.append([row['model'], row['requests'], row['total_tokens'], round(row['cost'], 2), req_pct, tok_pct, cost_pct])

bar_model = BarChart()
bar_model.type = 'col'
bar_model.title = 'Model Comparison (All Metrics)'
bar_model.y_axis.title = 'Value'
bar_model.x_axis.title = 'Model'
bar_model.grouping = 'clustered'
data = Reference(ws2b, min_col=2, min_row=1, max_row=len(model_summary)+1, max_col=4)
cats = Reference(ws2b, min_col=1, min_row=2, max_row=len(model_summary)+1)
bar_model.add_data(data, titles_from_data=True)
bar_model.set_categories(cats)
bar_model.height = 12
bar_model.width = 20
ws2b.add_chart(bar_model, 'J1')

# Sheet3: By Date - Combined 3 Metrics (Cost, Tokens, Requests)
date_summary = df.groupby('date').agg({
    'requests': 'sum',
    'total_tokens': 'sum',
    'cost': 'sum'
}).reset_index().sort_values('date')

ws3 = wb.create_sheet('ByDate_All')
# Original data
ws3.append(['Date', 'Cost(CNY)', 'TotalTokens', 'Requests', 'Cost(%)', 'Tokens(%)', 'Requests(%)'])
max_cost = date_summary['cost'].max()
max_tokens = date_summary['total_tokens'].max()
max_requests = date_summary['requests'].max()
for _, row in date_summary.iterrows():
    cost_pct = round(row['cost'] / max_cost * 100, 1)
    tokens_pct = round(row['total_tokens'] / max_tokens * 100, 1)
    requests_pct = round(row['requests'] / max_requests * 100, 1)
    ws3.append([row['date'].strftime('%Y-%m-%d'), round(row['cost'], 2), row['total_tokens'], row['requests'], cost_pct, tokens_pct, requests_pct])

# Create line chart with 3 metrics (normalized)
line = LineChart()
line.title = 'Daily Trends - All Metrics (Normalized)'
line.y_axis.title = 'Percentage of Max (%)'
line.x_axis.title = 'Date'

# Categories (dates)
cats = Reference(ws3, min_col=1, min_row=2, max_row=len(date_summary)+1)

# Data - three normalized series (columns E, F, G)
data = Reference(ws3, min_col=5, min_row=1, max_row=len(date_summary)+1, max_col=7)
line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.height = 12
line.width = 20

# Set different colors for each series
line.series[0].graphicalProperties.line.solidFill = "4472C4"  # Blue - Cost
line.series[1].graphicalProperties.line.solidFill = "ED7D31"  # Orange - Tokens
line.series[2].graphicalProperties.line.solidFill = "70AD47"  # Green - Requests

ws3.add_chart(line, 'J1')

# Sheet4: Daily Cost by Model (Stacked)
ws4 = wb.create_sheet('DailyByModel')
pivot = df.pivot_table(index='date', columns='model', values='cost', aggfunc='sum', fill_value=0).reset_index()
pivot['date'] = pivot['date'].dt.strftime('%Y-%m-%d')

headers = list(pivot.columns)
ws4.append(headers)
for row in pivot.values:
    ws4.append([round(x, 2) if isinstance(x, float) else x for x in row])

bar = BarChart()
bar.type = 'col'
bar.style = 10
bar.title = 'Daily Cost by Model (Stacked)'
bar.y_axis.title = 'Cost(CNY)'
bar.x_axis.title = 'Date'
bar.grouping = 'stacked'
bar.overlap = 100

data = Reference(ws4, min_col=2, min_row=1, max_row=len(pivot)+1, max_col=len(headers))
cats = Reference(ws4, min_col=1, min_row=2, max_row=len(pivot)+1)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.height = 12
bar.width = 20
ws4.add_chart(bar, 'J1')

# Sheet5: Token Analysis
ws5 = wb.create_sheet('TokenAnalysis')
token_summary = df.groupby('model').agg({
    'input_tokens': 'sum',
    'output_tokens': 'sum',
    'total_tokens': 'sum'
}).sort_values('total_tokens', ascending=False).reset_index()

ws5.append(['Model', 'InputTokens', 'OutputTokens', 'TotalTokens'])
for _, row in token_summary.iterrows():
    ws5.append([row['model'], row['input_tokens'], row['output_tokens'], row['total_tokens']])

ws5['F1'] = 'Model'
ws5['G1'] = 'InputPct'
ws5['H1'] = 'OutputPct'
for i, row in token_summary.iterrows():
    input_ratio = row['input_tokens'] / row['total_tokens'] * 100
    output_ratio = row['output_tokens'] / row['total_tokens'] * 100
    ws5[f'F{i+2}'] = row['model']
    ws5[f'G{i+2}'] = round(input_ratio, 1)
    ws5[f'H{i+2}'] = round(output_ratio, 1)

# Sheet6: Summary
ws6 = wb.create_sheet('Summary')
ws6['A1'] = 'Metric'
ws6['B1'] = 'Value'
ws6['A1'].font = Font(bold=True)
ws6['B1'].font = Font(bold=True)

daily_cost = df.groupby('date')['cost'].sum()
summary_data = [
    ['Total Cost (CNY)', round(df['cost'].sum(), 2)],
    ['Total Requests', df['requests'].sum()],
    ['Total Input Tokens', df['input_tokens'].sum()],
    ['Total Output Tokens', df['output_tokens'].sum()],
    ['Total Tokens', df['total_tokens'].sum()],
    ['Models Used', df['model'].nunique()],
    ['Days Used', df['date'].nunique()],
    ['Avg Daily Cost', round(daily_cost.mean(), 2)],
    ['Max Daily Cost', round(daily_cost.max(), 2)],
    ['Min Daily Cost', round(daily_cost.min(), 2)],
    ['Max Cost Date', daily_cost.idxmax().strftime('%Y-%m-%d')],
    ['Min Cost Date', daily_cost.idxmin().strftime('%Y-%m-%d')],
    ['Avg Cost per 1M Tokens', round(df['cost'].sum() / df['total_tokens'].sum() * 1000000, 2)],
    ['Avg Cost per Request', round(df['cost'].sum() / df['requests'].sum(), 4)],
]

for i, (key, value) in enumerate(summary_data, start=2):
    ws6[f'A{i}'] = key
    ws6[f'B{i}'] = value

for ws in [ws1, ws2, ws2b, ws3, ws4, ws5, ws6]:
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

wb.save('billing_analysis.xlsx')
print('Excel updated: billing_analysis.xlsx')
