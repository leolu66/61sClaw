#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报销单生成器
根据行程目录下的发票文件，智能生成报销单 Excel 并自动填入发票金额
"""

import json
import os
import re
import sys
import io
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass, field
from enum import Enum

# 设置 UTF-8 编码（Windows 兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# 添加脚本目录到路径
script_dir = Path(__file__).parent
sys.path.insert(0, str(script_dir))

from trip_parser import TripSegment


class InvoiceType(Enum):
    """发票类型"""
    TRAIN = "高铁/火车票"
    TAXI = "打车票"
    FLIGHT = "机票"
    HOTEL = "酒店"
    UNKNOWN = "未知"


@dataclass
class InvoiceInfo:
    """发票信息"""
    file_path: Path
    invoice_type: InvoiceType
    amount: float = 0.0
    date: Optional[str] = None
    departure: Optional[str] = None
    destination: Optional[str] = None
    invoice_no: Optional[str] = None
    matched: bool = False
    matched_segment: Optional[TripSegment] = None
    # 智能关联相关属性
    file_mtime: float = 0.0  # 文件修改时间
    related_invoices: List['InvoiceInfo'] = field(default_factory=list)  # 关联的相关票据


@dataclass
class ExpenseItem:
    """报销单条目"""
    sequence: int  # 序号
    route: str  # 路线描述
    departure_date: str  # 出发日期
    transport: str  # 交通工具
    amount: float  # 金额
    invoice_matched: bool = False  # 是否匹配到发票
    invoice_files: List[str] = field(default_factory=list)  # 匹配的发票文件


class InvoiceRecognizer:
    """发票识别器"""
    
    def __init__(self):
        self.ns = {
            'fp': 'http://www.chinatax.gov.cn/invoice'
        }
    
    def recognize_invoice(self, file_path: Path) -> Optional[InvoiceInfo]:
        """
        识别发票文件
        
        Args:
            file_path: 发票文件路径
            
        Returns:
            InvoiceInfo 或 None
        """
        suffix = file_path.suffix.lower()
        
        invoice_info = None
        if suffix == '.xml':
            invoice_info = self._parse_xml_invoice(file_path)
        elif suffix == '.pdf':
            invoice_info = self._parse_pdf_invoice(file_path)
        
        # 优先从文件名提取金额（更准确）
        filename_amount = self._extract_amount_from_filename(file_path.name)
        if invoice_info and filename_amount > 0:
            invoice_info.amount = filename_amount
        
        return invoice_info
    
    def _extract_amount_from_filename(self, filename: str) -> float:
        """从文件名中提取金额"""
        # 匹配模式: 优先匹配 日期_金额_ 格式 (如 260315_34.00_)
        patterns = [
            r'(\d+)_(\d+\.\d{2})_',   # 260315_34.00_北京滴滴
            r'_(\d+\.\d{2})_',       # _34.00_
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                try:
                    # 对于第一个模式，取第二个捕获组（金额）
                    if pattern.startswith(r'(\d+)'):
                        amount_str = match.group(2)
                    else:
                        amount_str = match.group(1)
                    return float(amount_str)
                except (ValueError, IndexError):
                    continue
        
        return 0.0
    
    def _parse_xml_invoice(self, file_path: Path) -> Optional[InvoiceInfo]:
        """解析 XML 格式发票"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # 尝试解析 XML
            try:
                root = ET.fromstring(content)
            except ET.ParseError:
                # 尝试清理后解析
                content = self._clean_xml_content(content)
                root = ET.fromstring(content)
            
            invoice_type = InvoiceType.UNKNOWN
            amount = 0.0
            date = None
            departure = None
            destination = None
            invoice_no = None
            
            # 尝试从 XML 中提取信息
            # 铁路电子客票
            if '铁路电子客票' in content or '火车票' in content:
                invoice_type = InvoiceType.TRAIN
                departure, destination = self._extract_train_route(content)
            # 旅客运输服务（打车）
            elif '旅客运输服务' in content or '出租车' in content or '网约车' in content:
                invoice_type = InvoiceType.TAXI
            # 机票
            elif '航空运输电子客票行程单' in content or '机票' in content or '航班' in content:
                invoice_type = InvoiceType.FLIGHT
                departure, destination = self._extract_flight_route(content)
            
            # 提取金额
            amount = self._extract_amount_from_xml(root, content)
            
            # 提取日期
            date = self._extract_date_from_xml(root, content)
            
            # 提取发票号码
            invoice_no = self._extract_invoice_no(content)
            
            return InvoiceInfo(
                file_path=file_path,
                invoice_type=invoice_type,
                amount=amount,
                date=date,
                departure=departure,
                destination=destination,
                invoice_no=invoice_no
            )
            
        except Exception as e:
            print(f"   ⚠️ 解析 XML 发票失败 {file_path.name}: {e}")
            return None
    
    def _clean_xml_content(self, content: str) -> str:
        """清理 XML 内容"""
        # 移除 BOM
        if content.startswith('\ufeff'):
            content = content[1:]
        # 修复常见的 XML 问题
        content = re.sub(r'xmlns="[^"]*"', '', content)
        return content
    
    def _parse_pdf_invoice(self, file_path: Path) -> Optional[InvoiceInfo]:
        """解析 PDF 格式发票（使用 OCR 或文本提取）"""
        try:
            # 首先根据文件名进行初步识别
            filename = file_path.name.lower()
            filename_type = InvoiceType.UNKNOWN
            
            if any(kw in filename for kw in ['滴滴', '出行', '打车', '久柏易游', '旅客运输']):
                filename_type = InvoiceType.TAXI
            elif any(kw in filename for kw in ['火车', '铁路', '高铁']):
                filename_type = InvoiceType.TRAIN
            elif any(kw in filename for kw in ['机票', '航空', '航班']):
                filename_type = InvoiceType.FLIGHT
            elif any(kw in filename for kw in ['酒店', '住宿']):
                filename_type = InvoiceType.HOTEL
            
            # 尝试使用 PyPDF2 提取文本
            try:
                import PyPDF2
                with open(file_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    text = ""
                    for page in reader.pages:
                        text += page.extract_text() or ""
            except Exception:
                text = ""
            
            # 如果 PyPDF2 失败或没有文本，尝试 OCR
            if not text.strip():
                text = self._ocr_pdf(file_path)
            
            invoice_type = InvoiceType.UNKNOWN
            amount = 0.0
            date = None
            departure = None
            destination = None
            invoice_no = None
            
            # 识别发票类型（优先使用文件名识别的结果）
            if filename_type != InvoiceType.UNKNOWN:
                invoice_type = filename_type
            elif '铁路电子客票' in text or '中国铁路' in text or ('G' in text and '次' in text):
                invoice_type = InvoiceType.TRAIN
                departure, destination = self._extract_train_route(text)
            elif '旅客运输服务' in text or '滴滴' in text or '出租车' in text or 'T3' in text or '曹操' in text:
                invoice_type = InvoiceType.TAXI
            elif '航空' in text or '航班' in text or '机票' in text or '登机牌' in text:
                invoice_type = InvoiceType.FLIGHT
                departure, destination = self._extract_flight_route(text)
            elif '酒店' in text or '住宿' in text or '宾馆' in text:
                invoice_type = InvoiceType.HOTEL
            
            # 提取金额
            amount = self._extract_amount_from_text(text)
            
            # 提取日期
            date = self._extract_date_from_text(text)
            
            # 提取发票号码
            invoice_no = self._extract_invoice_no(text)
            
            return InvoiceInfo(
                file_path=file_path,
                invoice_type=invoice_type,
                amount=amount,
                date=date,
                departure=departure,
                destination=destination,
                invoice_no=invoice_no
            )
            
        except Exception as e:
            print(f"   ⚠️ 解析 PDF 发票失败 {file_path.name}: {e}")
            return None
    
    def _ocr_pdf(self, file_path: Path) -> str:
        """使用 OCR 识别 PDF"""
        try:
            # 尝试使用 pdf2image + pytesseract
            from pdf2image import convert_from_path
            import pytesseract
            
            images = convert_from_path(str(file_path))
            text = ""
            for image in images:
                text += pytesseract.image_to_string(image, lang='chi_sim+eng')
            return text
        except Exception:
            return ""
    
    def _extract_amount_from_xml(self, root: ET.Element, content: str) -> float:
        """从 XML 中提取金额"""
        # 尝试多种金额字段
        amount_patterns = [
            r'价税合计.*?(\d+\.?\d*)',
            r'金额[：:]\s*(\d+\.?\d*)',
            r'¥\s*(\d+\.?\d*)',
        ]
        
        for pattern in amount_patterns:
            match = re.search(pattern, content)
            if match:
                try:
                    return float(match.group(1))
                except ValueError:
                    continue
        
        return 0.0
    
    def _extract_amount_from_text(self, text: str) -> float:
        """从文本中提取金额"""
        # 常见的金额模式
        patterns = [
            r'价税合计[：:]\s*\D*\s*(\d+\.\d{2})',
            r'合计金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'总金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥]\s*(\d+\.\d{2})',
            r'(\d+\.\d{2})\s*元',
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, text)
            if matches:
                # 返回最大的金额（通常是总计）
                try:
                    amounts = [float(m) for m in matches]
                    return max(amounts)
                except ValueError:
                    continue
        
        return 0.0
    
    def _extract_date_from_xml(self, root: ET.Element, content: str) -> Optional[str]:
        """从 XML 中提取日期"""
        patterns = [
            r'开票日期\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
            r'(\d{4}年\d{1,2}月\d{1,2}日)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, content)
            if match:
                return match.group(1)
        
        return None
    
    def _extract_date_from_text(self, text: str) -> Optional[str]:
        """从文本中提取日期"""
        patterns = [
            r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
            r'(\d{4}年\d{1,2}月\d{1,2}日)',
            r'日期[：:]\s*(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)
        
        return None
    
    def _extract_train_route(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        """提取火车行程路线"""
        # 匹配模式：出发站 → 到达站
        patterns = [
            r'([\u4e00-\u9fa5]+?站)\s*[-→~到至]\s*([\u4e00-\u9fa5]+?站)',
            r'出发站[：:]\s*([\u4e00-\u9fa5]+?站).*?到达站[：:]\s*([\u4e00-\u9fa5]+?站)',
            r'([\u4e00-\u9fa5]+?)\s*[-→~到至]\s*([\u4e00-\u9fa5]+?)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                departure = match.group(1).replace('站', '')
                destination = match.group(2).replace('站', '')
                return departure, destination
        
        return None, None
    
    def _extract_flight_route(self, text: str) -> Tuple[Optional[str], Optional[str]]:
        """提取飞机行程路线"""
        patterns = [
            r'([\u4e00-\u9fa5]+?)\s*[-→~到至]\s*([\u4e00-\u9fa5]+?)',
            r'出发[：:]\s*([\u4e00-\u9fa5]+).*?到达[：:]\s*([\u4e00-\u9fa5]+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1), match.group(2)
        
        return None, None
    
    def _extract_invoice_no(self, text: str) -> Optional[str]:
        """提取发票号码"""
        patterns = [
            r'发票号码[：:]\s*(\d+)',
            r'发票代码[：:]\s*(\d+)',
            r'No[.:]\s*(\d+)',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return match.group(1)
        
        return None


class ExpenseReportGenerator:
    """报销单生成器"""
    
    def __init__(self, config: Dict):
        self.config = config
        self.recognizer = InvoiceRecognizer()
        self.invoices: List[InvoiceInfo] = []
        self.segments: List[TripSegment] = []
    
    def scan_invoices(self, trip_dir: Path) -> List[InvoiceInfo]:
        """
        扫描行程目录下的所有发票文件（只扫描PDF，跳过XML）
        
        Args:
            trip_dir: 行程目录路径
            
        Returns:
            发票信息列表
        """
        invoices = []
        
        # 扫描所有PDF文件（跳过XML，因为内容相同）
        for file_path in trip_dir.rglob('*.pdf'):
            if file_path.is_file():
                invoice = self.recognizer.recognize_invoice(file_path)
                if invoice:
                    invoices.append(invoice)
                    print(f"   ✅ 识别发票: {file_path.name} - {invoice.invoice_type.value} - ¥{invoice.amount:.2f}")
        
        return invoices
    
    def match_invoices_to_trip(self, invoices: List[InvoiceInfo], trip_info: Dict) -> None:
        """
        根据 trip_info.json 的路段信息按顺序匹配发票
        
        匹配策略：
        1. 按 outward（去程）顺序匹配路段
        2. 按 return（回程）顺序匹配路段
        3. 每段根据 transport 类型和地点匹配对应发票
        
        Args:
            invoices: 发票列表
            trip_info: 行程信息字典
        """
        departure_city = trip_info.get('departure_city', '')
        destination_city = trip_info.get('destination_city', '')
        start_date = trip_info.get('start_date', '')
        end_date = trip_info.get('end_date', '')
        outward_stages = trip_info.get('outward', [])
        return_stages = trip_info.get('return', [])
        
        print(f"   📍 行程: {departure_city} → {destination_city}")
        
        # 第一步：智能关联相关票据
        self._associate_related_invoices(invoices)
        
        # 第二步：按路段顺序匹配发票
        if outward_stages and return_stages:
            # 有详细路段信息
            print(f"   📝 去程 {len(outward_stages)} 段, 回程 {len(return_stages)} 段")
            
            # 去程匹配
            print("   🚗 匹配去程路段...")
            for i, stage in enumerate(outward_stages, 1):
                self._match_stage_to_invoice(stage, invoices, departure_city, destination_city, f"去程-{i}", start_date)
            
            # 回程匹配
            print("   🚗 匹配回程路段...")
            for i, stage in enumerate(return_stages, 1):
                self._match_stage_to_invoice(stage, invoices, departure_city, destination_city, f"回程-{i}", end_date)
        else:
            # 只有基本信息，简化匹配
            print("   📝 简化模式：匹配高铁和打车票")
            
            # 匹配火车票（去程）
            for invoice in invoices:
                if invoice.matched:
                    continue
                if invoice.invoice_type == InvoiceType.TRAIN:
                    invoice.matched = True
                    invoice.trip_stage = "去程-高铁"
                    print(f"      ✓ 去程高铁: {departure_city} → {destination_city} ← {invoice.file_path.name}")
                    break
            
            # 匹配打车票（按日期和城市）
            for invoice in invoices:
                if invoice.matched:
                    continue
                if invoice.invoice_type == InvoiceType.TAXI:
                    # 简化匹配：未匹配的打车票都标记为已匹配
                    invoice.matched = True
                    print(f"      ✓ 市内交通: {invoice.file_path.name}")
        
        # 统计匹配结果
        matched_count = len([inv for inv in invoices if inv.matched])
        print(f"   ✅ 成功匹配 {matched_count}/{len(invoices)} 张发票")
    
    def _match_stage_to_invoice(self, stage: Dict, invoices: List[InvoiceInfo], 
                                 departure_city: str, destination_city: str, stage_name: str,
                                 stage_date: Optional[str] = None) -> None:
        """
        将单个路段匹配到发票
        
        Args:
            stage: 路段信息 {stage, from, to, transport}
            invoices: 发票列表
            departure_city: 出发城市
            destination_city: 目的城市
            stage_name: 路段名称（用于日志）
            stage_date: 路段日期（可选，用于日期匹配）
        """
        from_loc = stage.get('from', '')
        to_loc = stage.get('to', '')
        transport = stage.get('transport', '')
        
        # 确定城市
        stage_city = self._determine_stage_city(from_loc, to_loc, departure_city, destination_city)
        
        # 根据 transport 确定需要的发票类型
        required_type = self._transport_to_invoice_type(transport)
        
        if required_type is None:
            return  # 不需要发票的路段（如地铁）
        
        # 首先尝试根据日期匹配（如果提供了路段日期）
        if stage_date:
            for invoice in invoices:
                if invoice.matched:
                    continue
                
                # 从文件名提取日期
                invoice_date_prefix = self._extract_date_prefix_from_filename(invoice.file_path.name)
                if invoice_date_prefix:
                    # 将路段日期转换为 YYMMDD 格式进行比较
                    try:
                        stage_dt = datetime.strptime(stage_date, '%Y-%m-%d')
                        stage_prefix = stage_dt.strftime('%y%m%d')
                        
                        # 如果日期匹配，优先使用
                        if invoice_date_prefix == stage_prefix:
                            # 检查类型和城市
                            if invoice.invoice_type == required_type:
                                if self._is_invoice_for_city(invoice, stage_city, departure_city, destination_city):
                                    invoice.matched = True
                                    invoice.trip_stage = stage_name
                                    invoice.stage_info = stage
                                    print(f"      ✓ {stage_name}: {from_loc} → {to_loc} ({transport}) ← {invoice.file_path.name} [日期匹配]")
                                    return
                    except ValueError:
                        pass
        
        # 日期不匹配或没有日期，按原来的逻辑匹配
        for invoice in invoices:
            if invoice.matched:
                continue
            
            # 检查发票类型是否匹配
            if invoice.invoice_type == required_type:
                # 检查城市是否匹配（通过文件名）
                if self._is_invoice_for_city(invoice, stage_city, departure_city, destination_city):
                    invoice.matched = True
                    invoice.trip_stage = stage_name
                    invoice.stage_info = stage
                    print(f"      ✓ {stage_name}: {from_loc} → {to_loc} ({transport}) ← {invoice.file_path.name}")
                    return
        
        # 没有找到匹配的发票
        print(f"      ⚠ {stage_name}: {from_loc} → {to_loc} ({transport}) - 未找到匹配发票")
    
    def _determine_stage_city(self, from_loc: str, to_loc: str, departure_city: str, destination_city: str) -> str:
        """确定路段所属城市"""
        # 根据地点关键词判断
        departure_keywords = ['家', '地铁站', '南京南站', '南京站', '南京']
        destination_keywords = ['宿舍', '草桥站', '北京南站', '北京站', '北京']
        
        if any(kw in from_loc for kw in departure_keywords) or any(kw in to_loc for kw in departure_keywords):
            return departure_city
        elif any(kw in from_loc for kw in destination_keywords) or any(kw in to_loc for kw in destination_keywords):
            return destination_city
        
        return departure_city  # 默认返回出发城市
    
    def _transport_to_invoice_type(self, transport: str) -> Optional[InvoiceType]:
        """将 transport 转换为发票类型"""
        transport_map = {
            '高铁': InvoiceType.TRAIN,
            '火车': InvoiceType.TRAIN,
            '打车': InvoiceType.TAXI,
            '出租车': InvoiceType.TAXI,
            '网约车': InvoiceType.TAXI,
            '地铁': InvoiceType.TAXI,  # 地铁行程单也作为打车票处理
            '机场大巴': InvoiceType.TAXI,
        }
        return transport_map.get(transport, InvoiceType.TAXI)  # 默认打车票
    
    def _is_invoice_for_city(self, invoice: InvoiceInfo, city: str, departure_city: str, destination_city: str) -> bool:
        """判断发票是否属于指定城市"""
        filename = invoice.file_path.name.lower()
        
        # 火车票特殊处理 - 不检查城市，直接匹配
        if invoice.invoice_type == InvoiceType.TRAIN:
            return True
        
        # 检查发票文件名
        if city == departure_city:
            # 南京相关关键词
            if any(kw in filename for kw in ['南京', 'nanjing', 'nj']):
                return True
        elif city == destination_city:
            # 北京相关关键词
            if any(kw in filename for kw in ['北京', 'beijing', 'bj']):
                return True
        
        # 如果发票文件名没有城市信息，检查关联的行程单
        related_invoices = getattr(invoice, 'related_invoices', [])
        for related in related_invoices:
            related_filename = related.file_path.name.lower()
            if city == departure_city:
                if any(kw in related_filename for kw in ['南京', 'nanjing', 'nj']):
                    return True
            elif city == destination_city:
                if any(kw in related_filename for kw in ['北京', 'beijing', 'bj']):
                    return True
        
        # 如果城市无法判断，对于打车票默认允许匹配（后续可根据金额等进一步筛选）
        if invoice.invoice_type == InvoiceType.TAXI:
            # 检查是否已经被其他路段匹配（通过检查是否所有关联票据都没有城市信息）
            has_city_info = False
            if any(kw in filename for kw in ['南京', '北京', 'nanjing', 'beijing', 'nj', 'bj']):
                has_city_info = True
            for related in related_invoices:
                if any(kw in related.file_path.name.lower() for kw in ['南京', '北京', 'nanjing', 'beijing', 'nj', 'bj']):
                    has_city_info = True
            
            # 如果没有城市信息，默认可以匹配任何城市（让后续逻辑决定具体匹配哪一段）
            if not has_city_info:
                return True
        
        return False
    
    def _associate_related_invoices(self, invoices: List[InvoiceInfo]) -> None:
        """
        智能关联相关票据
        
        关联规则：
        1. 文件名中包含相同日期前缀（如 260315）和相同金额，一个是主发票，另一个是行程单
        2. 文件名为 invoice.pdf 和 trip.pdf（英文含义：发票和行程单）
        3. 金额相同且文件时间相近
        
        Args:
            invoices: 发票列表
        """
        if len(invoices) < 2:
            return
        
        print("   🔗 智能关联相关票据...")
        
        # 获取文件修改时间
        for invoice in invoices:
            try:
                stat = invoice.file_path.stat()
                invoice.file_mtime = stat.st_mtime
            except Exception:
                invoice.file_mtime = 0
        
        associated_count = 0
        
        # 规则1：按日期前缀+金额分组关联
        date_amount_groups: Dict[str, List[InvoiceInfo]] = {}
        for invoice in invoices:
            if invoice.amount <= 0:
                continue
            
            filename = invoice.file_path.name
            date_prefix = self._extract_date_prefix_from_filename(filename)
            
            if date_prefix:
                key = f"{date_prefix}_{invoice.amount:.2f}"
                if key not in date_amount_groups:
                    date_amount_groups[key] = []
                date_amount_groups[key].append(invoice)
        
        for key, group in date_amount_groups.items():
            if len(group) < 2:
                continue
            
            main_invoice = None
            itinerary = None
            
            for inv in group:
                filename = inv.file_path.name.lower()
                if '行程单' in filename:
                    itinerary = inv
                else:
                    main_invoice = inv
            
            if main_invoice and itinerary:
                self._link_invoices(main_invoice, itinerary)
                associated_count += 1
                print(f"      ✓ 关联: {main_invoice.file_path.name} ↔ {itinerary.file_path.name} (金额: ¥{main_invoice.amount:.2f})")
        
        # 规则2：invoice.pdf 和 trip.pdf 关联
        invoice_pdf = None
        trip_pdf = None
        
        for inv in invoices:
            filename_lower = inv.file_path.name.lower()
            if filename_lower == 'invoice.pdf':
                invoice_pdf = inv
            elif filename_lower == 'trip.pdf':
                trip_pdf = inv
        
        if invoice_pdf and trip_pdf:
            # 检查金额是否相同或相近
            if abs(invoice_pdf.amount - trip_pdf.amount) < 0.01:
                self._link_invoices(invoice_pdf, trip_pdf)
                associated_count += 1
                print(f"      ✓ 关联: {invoice_pdf.file_path.name} ↔ {trip_pdf.file_path.name} (金额: ¥{invoice_pdf.amount:.2f})")
        
        if associated_count > 0:
            print(f"   ✅ 发现 {associated_count} 组相关票据")
    
    def _link_invoices(self, inv1: InvoiceInfo, inv2: InvoiceInfo) -> None:
        """建立两个发票之间的关联关系"""
        if not hasattr(inv1, 'related_invoices'):
            inv1.related_invoices = []
        if not hasattr(inv2, 'related_invoices'):
            inv2.related_invoices = []
        
        if inv2 not in inv1.related_invoices:
            inv1.related_invoices.append(inv2)
        if inv1 not in inv2.related_invoices:
            inv2.related_invoices.append(inv1)
    
    def _extract_date_prefix_from_filename(self, filename: str) -> Optional[str]:
        """从文件名中提取日期前缀（如 260315）"""
        # 匹配模式: 260315_34.00_ 中的 260315
        match = re.search(r'^(\d{6})_', filename)
        if match:
            return match.group(1)
        return None
    
    def _calculate_match_score(self, invoice: InvoiceInfo, segment: TripSegment) -> float:
        """计算发票与行程的匹配分数"""
        score = 0.0
        
        # 城市匹配
        if invoice.departure and invoice.destination:
            dep_match = invoice.departure in segment.departure_city or segment.departure_city in invoice.departure
            dest_match = invoice.destination in segment.destination_city or segment.destination_city in invoice.destination
            
            if dep_match and dest_match:
                score += 0.5
            elif dep_match or dest_match:
                score += 0.3
        
        # 日期匹配
        if invoice.date:
            try:
                # 解析发票日期
                invoice_date = self._parse_date(invoice.date)
                if invoice_date:
                    if segment.start_date <= invoice_date <= segment.end_date:
                        score += 0.5
                    # 允许前后一天的误差
                    elif abs((invoice_date - segment.start_date).days) <= 1:
                        score += 0.3
            except Exception:
                pass
        
        return min(score, 1.0)
    
    def _parse_date(self, date_str: str) -> Optional[datetime]:
        """解析日期字符串"""
        formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y年%m月%d日',
            '%m-%d',
            '%m/%d',
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        return None
    
    def generate_expense_items_from_trip(self, invoices: List[InvoiceInfo], trip_info: Dict) -> List[ExpenseItem]:
        """
        根据 trip_info 生成报销单条目
        
        支持两种模式：
        1. 有详细路段信息（outward/return）：按路段生成条目
        2. 只有基本信息：生成简化的去程/回程条目
        
        Args:
            invoices: 发票列表
            trip_info: 行程信息
            
        Returns:
            报销单条目列表
        """
        items = []
        departure_city = trip_info.get('departure_city', '')
        destination_city = trip_info.get('destination_city', '')
        start_date = trip_info.get('start_date', '')
        end_date = trip_info.get('end_date', '')
        transport = trip_info.get('transport', '火车')
        outward_stages = trip_info.get('outward', [])
        return_stages = trip_info.get('return', [])
        
        def get_invoice_files_with_related(invoice: InvoiceInfo) -> List[str]:
            """获取发票文件及其关联票据的文件名"""
            if not invoice:
                return []
            files = [str(invoice.file_path.name)]
            for related in getattr(invoice, 'related_invoices', []):
                files.append(str(related.file_path.name))
            return files
        
        def find_matched_invoice(stage_name: str) -> Optional[InvoiceInfo]:
            """查找匹配到指定路段的发票"""
            for inv in invoices:
                if inv.matched and getattr(inv, 'trip_stage', '') == stage_name:
                    return inv
            return None
        
        # 判断是否有详细路段信息
        if outward_stages and return_stages:
            # 模式1：有详细路段信息，按路段生成
            # ========== 去程 ==========
            print("   📝 生成去程报销条目...")
            for i, stage in enumerate(outward_stages, 1):
                stage_name = f"去程-{i}"
                invoice = find_matched_invoice(stage_name)
                
                item = ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{stage.get('from', '')} → {stage.get('to', '')}",
                    departure_date=start_date,
                    transport=stage.get('transport', ''),
                    amount=invoice.amount if invoice else 0.0,
                    invoice_matched=invoice is not None,
                    invoice_files=get_invoice_files_with_related(invoice) if invoice else []
                )
                items.append(item)
                if invoice:
                    print(f"      ✓ {stage_name}: {item.route} ← ¥{invoice.amount:.2f}")
                else:
                    print(f"      ⚠ {stage_name}: {item.route} - 无发票")
            
            # ========== 回程 ==========
            print("   📝 生成回程报销条目...")
            for i, stage in enumerate(return_stages, 1):
                stage_name = f"回程-{i}"
                invoice = find_matched_invoice(stage_name)
                
                item = ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{stage.get('from', '')} → {stage.get('to', '')}",
                    departure_date=end_date,
                    transport=stage.get('transport', ''),
                    amount=invoice.amount if invoice else 0.0,
                    invoice_matched=invoice is not None,
                    invoice_files=get_invoice_files_with_related(invoice) if invoice else []
                )
                items.append(item)
                if invoice:
                    print(f"      ✓ {stage_name}: {item.route} ← ¥{invoice.amount:.2f}")
                else:
                    print(f"      ⚠ {stage_name}: {item.route} - 无发票")
        else:
            # 模式2：只有基本信息，生成简化条目
            print("   📝 生成简化报销条目...")
            
            # 去程高铁
            train_invoice = find_matched_invoice("去程-高铁")
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{departure_city} → {destination_city}",
                departure_date=start_date,
                transport=transport,
                amount=train_invoice.amount if train_invoice else 0.0,
                invoice_matched=train_invoice is not None,
                invoice_files=get_invoice_files_with_related(train_invoice) if train_invoice else []
            ))
            
            # 其他发票按顺序添加
            other_invoices = [inv for inv in invoices if inv.matched and inv.invoice_type == InvoiceType.TAXI]
            for i, invoice in enumerate(other_invoices, 2):
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{departure_city}市内交通" if i <= len(other_invoices)//2 + 1 else f"{destination_city}市内交通",
                    departure_date=start_date if i <= len(other_invoices)//2 + 1 else end_date,
                    transport="打车/地铁",
                    amount=invoice.amount,
                    invoice_matched=True,
                    invoice_files=get_invoice_files_with_related(invoice)
                ))
        
        return items
    
    def _create_train_items(self, segment: TripSegment, train_invoices: List[InvoiceInfo]) -> List[ExpenseItem]:
        """创建高铁/火车报销条目"""
        items = []
        
        # 去程
        if train_invoices:
            invoice = train_invoices.pop(0)
            # 收集关联票据的文件名
            invoice_files = [str(invoice.file_path.name)]
            for related in getattr(invoice, 'related_invoices', []):
                invoice_files.append(str(related.file_path.name))
            
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city} → {segment.destination_city}",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport="高铁/火车",
                amount=invoice.amount,
                invoice_matched=True,
                invoice_files=invoice_files
            ))
        else:
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city} → {segment.destination_city}",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport="高铁/火车",
                amount=0.0,
                invoice_matched=False,
                invoice_files=[]
            ))
        
        # 回程（如果行程超过1天）
        if (segment.end_date - segment.start_date).days > 0:
            if train_invoices:
                invoice = train_invoices.pop(0)
                # 收集关联票据的文件名
                invoice_files = [str(invoice.file_path.name)]
                for related in getattr(invoice, 'related_invoices', []):
                    invoice_files.append(str(related.file_path.name))
                
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{segment.destination_city} → {segment.departure_city}",
                    departure_date=segment.end_date.strftime('%Y-%m-%d'),
                    transport="高铁/火车",
                    amount=invoice.amount,
                    invoice_matched=True,
                    invoice_files=invoice_files
                ))
            else:
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{segment.destination_city} → {segment.departure_city}",
                    departure_date=segment.end_date.strftime('%Y-%m-%d'),
                    transport="高铁/火车",
                    amount=0.0,
                    invoice_matched=False,
                    invoice_files=[]
                ))
        
        return items
    
    def _create_flight_items(self, segment: TripSegment, flight_invoices: List[InvoiceInfo], taxi_invoices: List[InvoiceInfo]) -> List[ExpenseItem]:
        """创建飞机报销条目"""
        items = []

        def get_invoice_files_with_related(invoice: InvoiceInfo) -> List[str]:
            """获取发票文件及其关联票据的文件名"""
            files = [str(invoice.file_path.name)]
            for related in getattr(invoice, 'related_invoices', []):
                files.append(str(related.file_path.name))
            return files

        # 去程 - 市区到机场（打车）
        taxi_to_airport = [inv for inv in taxi_invoices if self._is_departure_taxi(inv, segment)]
        if taxi_to_airport:
            invoice = taxi_to_airport.pop(0)
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city}市区 → {segment.departure_city}机场",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport="打车",
                amount=invoice.amount,
                invoice_matched=True,
                invoice_files=get_invoice_files_with_related(invoice)
            ))

        # 去程 - 机票
        if flight_invoices:
            invoice = flight_invoices.pop(0)
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city} → {segment.destination_city}",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport="飞机",
                amount=invoice.amount,
                invoice_matched=True,
                invoice_files=get_invoice_files_with_related(invoice)
            ))
        else:
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city} → {segment.destination_city}",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport="飞机",
                amount=0.0,
                invoice_matched=False,
                invoice_files=[]
            ))

        # 去程 - 机场到市区（打车）
        taxi_from_airport = [inv for inv in taxi_invoices if self._is_arrival_taxi(inv, segment)]
        if taxi_from_airport:
            invoice = taxi_from_airport.pop(0)
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.destination_city}机场 → {segment.destination_city}市区",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport="打车",
                amount=invoice.amount,
                invoice_matched=True,
                invoice_files=get_invoice_files_with_related(invoice)
            ))

        # 回程（如果行程超过1天）
        if (segment.end_date - segment.start_date).days > 0:
            # 回程 - 市区到机场（打车）
            if taxi_from_airport:
                invoice = taxi_from_airport.pop(0)
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{segment.destination_city}市区 → {segment.destination_city}机场",
                    departure_date=segment.end_date.strftime('%Y-%m-%d'),
                    transport="打车",
                    amount=invoice.amount,
                    invoice_matched=True,
                    invoice_files=get_invoice_files_with_related(invoice)
                ))

            # 回程 - 机票
            if flight_invoices:
                invoice = flight_invoices.pop(0)
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{segment.destination_city} → {segment.departure_city}",
                    departure_date=segment.end_date.strftime('%Y-%m-%d'),
                    transport="飞机",
                    amount=invoice.amount,
                    invoice_matched=True,
                    invoice_files=get_invoice_files_with_related(invoice)
                ))
            else:
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{segment.destination_city} → {segment.departure_city}",
                    departure_date=segment.end_date.strftime('%Y-%m-%d'),
                    transport="飞机",
                    amount=0.0,
                    invoice_matched=False,
                    invoice_files=[]
                ))

            # 回程 - 机场到市区（打车）
            if taxi_to_airport:
                invoice = taxi_to_airport.pop(0)
                items.append(ExpenseItem(
                    sequence=len(items) + 1,
                    route=f"{segment.departure_city}机场 → {segment.departure_city}市区",
                    departure_date=segment.end_date.strftime('%Y-%m-%d'),
                    transport="打车",
                    amount=invoice.amount,
                    invoice_matched=True,
                    invoice_files=get_invoice_files_with_related(invoice)
                ))

        return items
    
    def _create_generic_items(self, segment: TripSegment, invoices: List[InvoiceInfo]) -> List[ExpenseItem]:
        """创建通用报销条目"""
        items = []

        def get_invoice_files_with_related(invoice: InvoiceInfo) -> List[str]:
            """获取发票文件及其关联票据的文件名"""
            files = [str(invoice.file_path.name)]
            for related in getattr(invoice, 'related_invoices', []):
                files.append(str(related.file_path.name))
            return files

        # 去程
        matched = [inv for inv in invoices if inv.matched]
        if matched:
            invoice = matched.pop(0)
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city} → {segment.destination_city}",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport=segment.transport_mode or "其他",
                amount=invoice.amount,
                invoice_matched=True,
                invoice_files=get_invoice_files_with_related(invoice)
            ))
        else:
            items.append(ExpenseItem(
                sequence=len(items) + 1,
                route=f"{segment.departure_city} → {segment.destination_city}",
                departure_date=segment.start_date.strftime('%Y-%m-%d'),
                transport=segment.transport_mode or "其他",
                amount=0.0,
                invoice_matched=False,
                invoice_files=[]
            ))

        return items
    
    def _is_departure_taxi(self, invoice: InvoiceInfo, segment: TripSegment) -> bool:
        """判断是否为出发地打车"""
        if not invoice.date:
            return True  # 无法判断时默认包含
        try:
            invoice_date = self._parse_date(invoice.date)
            if invoice_date:
                return abs((invoice_date - segment.start_date).days) <= 1
        except Exception:
            pass
        return True
    
    def _is_arrival_taxi(self, invoice: InvoiceInfo, segment: TripSegment) -> bool:
        """判断是否为到达地打车"""
        if not invoice.date:
            return True
        try:
            invoice_date = self._parse_date(invoice.date)
            if invoice_date:
                return abs((invoice_date - segment.start_date).days) <= 1
        except Exception:
            pass
        return True
    
    def generate_excel(self, trip_dir: Path, segments: List[TripSegment], items: List[ExpenseItem]) -> Path:
        """
        生成报销单 Excel
        
        Args:
            trip_dir: 行程目录
            segments: 行程段列表
            items: 报销单条目列表
            
        Returns:
            生成的 Excel 文件路径
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            print("   ⚠️ 未安装 openpyxl，尝试安装...")
            os.system(f"{sys.executable} -m pip install openpyxl -q")
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "差旅报销单"
        
        # 设置标题
        ws['A1'] = "差旅报销单"
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置表头
        headers = ["序号", "路线", "出发日期", "返回日期", "交通工具", "金额(元)", "发票匹配", "发票文件"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for row, item in enumerate(items, 4):
            ws.cell(row=row, column=1, value=item.sequence)
            ws.cell(row=row, column=2, value=item.route)
            ws.cell(row=row, column=3, value=item.departure_date)
            # 返回日期列已删除
            ws.cell(row=row, column=5, value=item.transport)
            ws.cell(row=row, column=6, value=item.amount)
            ws.cell(row=row, column=7, value="是" if item.invoice_matched else "否")
            ws.cell(row=row, column=8, value=", ".join(item.invoice_files))
            
            # 设置金额格式
            ws.cell(row=row, column=6).number_format = '0.00'
        
        # 添加合计行
        total_row = len(items) + 4
        ws.cell(row=total_row, column=1, value="合计")
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        total_amount = sum(item.amount for item in items)
        ws.cell(row=total_row, column=6, value=total_amount)
        ws.cell(row=total_row, column=6).number_format = '0.00'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 40
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(3, total_row + 1):
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border
        
        # 保存文件
        timestamp = datetime.now().strftime('%H%M%S')
        output_file = trip_dir / f"报销单_{timestamp}.xlsx"
        wb.save(str(output_file))
        
        return output_file
    
    def generate_match_report(self, trip_dir: Path, invoices: List[InvoiceInfo]) -> Path:
        """
        生成发票匹配报告
        
        Args:
            trip_dir: 行程目录
            invoices: 发票列表
            
        Returns:
            报告文件路径
        """
        report_file = trip_dir / "发票匹配报告.txt"
        
        matched = [inv for inv in invoices if inv.matched]
        unmatched = [inv for inv in invoices if not inv.matched]
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("发票匹配报告\n")
            f.write("=" * 60 + "\n\n")
            
            f.write(f"扫描时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"发票总数: {len(invoices)}\n")
            f.write(f"已匹配: {len(matched)}\n")
            f.write(f"未匹配: {len(unmatched)}\n\n")
            
            # 已匹配发票
            if matched:
                f.write("-" * 60 + "\n")
                f.write("已匹配发票:\n")
                f.write("-" * 60 + "\n")
                for inv in matched:
                    f.write(f"\n文件: {inv.file_path.name}\n")
                    f.write(f"  类型: {inv.invoice_type.value}\n")
                    f.write(f"  金额: ¥{inv.amount:.2f}\n")
                    if inv.departure and inv.destination:
                        f.write(f"  路线: {inv.departure} → {inv.destination}\n")
                    if inv.date:
                        f.write(f"  日期: {inv.date}\n")
                    if inv.invoice_no:
                        f.write(f"  发票号: {inv.invoice_no}\n")
                    # 显示关联票据
                    related = getattr(inv, 'related_invoices', [])
                    if related:
                        f.write(f"  关联票据: {', '.join([r.file_path.name for r in related])}\n")
            
            # 未匹配发票
            if unmatched:
                f.write("\n" + "-" * 60 + "\n")
                f.write("未匹配发票:\n")
                f.write("-" * 60 + "\n")
                for inv in unmatched:
                    f.write(f"\n文件: {inv.file_path.name}\n")
                    f.write(f"  类型: {inv.invoice_type.value}\n")
                    f.write(f"  金额: ¥{inv.amount:.2f}\n")
                    if inv.departure and inv.destination:
                        f.write(f"  路线: {inv.departure} → {inv.destination}\n")
                    if inv.date:
                        f.write(f"  日期: {inv.date}\n")
                    # 显示关联票据
                    related = getattr(inv, 'related_invoices', [])
                    if related:
                        f.write(f"  关联票据: {', '.join([r.file_path.name for r in related])}\n")
        
        return report_file


def load_config() -> Dict:
    """加载配置文件"""
    config_path = Path(__file__).parent / "config.json"
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"⚠️ 无法加载配置文件: {e}")
        print("使用默认配置")
        return {
            "base_path": "D:\\Users\\luzhe\\报销凭证",
            "expense_report": {
                "auto_generate": True,
                "match_threshold": 0.5,
                "output_filename": "报销单.xlsx",
                "report_filename": "发票匹配报告.txt"
            }
        }


def generate_expense_report(trip_dir: Path, segments: List[TripSegment]) -> Dict:
    """
    生成报销单主函数
    
    Args:
        trip_dir: 行程目录路径
        segments: 行程段列表
        
    Returns:
        执行结果
    """
    print(f"\n📊 开始生成报销单...")
    print(f"   行程目录: {trip_dir}")
    
    # 加载配置
    config = load_config()
    
    # 初始化生成器
    generator = ExpenseReportGenerator(config)
    
    # 扫描发票（只扫描PDF）
    print("\n🔍 扫描发票文件...")
    invoices = generator.scan_invoices(trip_dir)
    print(f"✅ 共识别 {len(invoices)} 张发票")
    
    if not invoices:
        print("⚠️ 未找到可识别的发票文件")
        return {
            "success": False,
            "error": "未找到发票文件",
            "excel_file": None,
            "report_file": None
        }
    
    # 尝试读取 trip_info.json
    trip_info_path = trip_dir / "trip_info.json"
    trip_info = None
    if trip_info_path.exists():
        try:
            with open(trip_info_path, 'r', encoding='utf-8') as f:
                trip_info = json.load(f)
            print(f"\n📋 读取行程信息: {trip_info_path.name}")
        except Exception as e:
            print(f"⚠️ 读取 trip_info.json 失败: {e}")
    
    # 匹配发票到行程
    print("\n🔗 匹配发票到行程...")
    if trip_info:
        # 使用新的基于 trip_info 的匹配逻辑
        generator.match_invoices_to_trip(invoices, trip_info)
    else:
        # 回退到旧的匹配逻辑
        generator.match_invoices_to_segments(invoices, segments)
    
    matched_count = len([inv for inv in invoices if inv.matched])
    print(f"✅ 成功匹配 {matched_count}/{len(invoices)} 张发票")
    
    # 生成报销单条目
    print("\n📝 生成报销单条目...")
    all_items = []
    
    if trip_info:
        # 使用 trip_info 生成条目
        all_items = generator.generate_expense_items_from_trip(invoices, trip_info)
    else:
        # 回退到旧的生成逻辑
        for segment in segments:
            items = generator.generate_expense_items(segment, invoices)
            all_items.extend(items)
    
    # 重新编号
    for i, item in enumerate(all_items, 1):
        item.sequence = i
    
    print(f"✅ 生成 {len(all_items)} 条报销记录")
    
    # 生成 Excel
    print("\n📄 生成报销单 Excel...")
    excel_file = generator.generate_excel(trip_dir, segments, all_items)
    print(f"✅ 报销单已保存: {excel_file}")
    
    # 生成匹配报告
    print("\n📋 生成发票匹配报告...")
    report_file = generator.generate_match_report(trip_dir, invoices)
    print(f"✅ 匹配报告已保存: {report_file}")
    
    # 计算总金额
    total_amount = sum(item.amount for item in all_items)
    matched_amount = sum(inv.amount for inv in invoices if inv.matched)
    
    print("\n" + "=" * 50)
    print("📊 报销单生成完成")
    print("=" * 50)
    print(f"报销条目数: {len(all_items)}")
    print(f"发票总数: {len(invoices)}")
    print(f"已匹配: {matched_count}")
    print(f"未匹配: {len(invoices) - matched_count}")
    print(f"报销总金额: ¥{total_amount:.2f}")
    print(f"发票匹配金额: ¥{matched_amount:.2f}")
    print("=" * 50)
    
    return {
        "success": True,
        "excel_file": str(excel_file),
        "report_file": str(report_file),
        "total_items": len(all_items),
        "total_invoices": len(invoices),
        "matched_invoices": matched_count,
        "total_amount": total_amount,
        "matched_amount": matched_amount
    }


def main():
    """主函数 - 命令行入口"""
    print("🚀 差旅报销单生成器")
    print("=" * 50)
    
    # 解析命令行参数
    if len(sys.argv) < 2:
        print("\n使用方法:")
        print(f"  python {sys.argv[0]} <行程目录路径>")
        print(f"  python {sys.argv[0]} --trip-dir <路径>")
        print("\n示例:")
        print(f'  python {sys.argv[0]} "D:\\Users\\luzhe\\报销凭证\\南京-北京-0208-0316"')
        sys.exit(1)
    
    # 获取行程目录
    if sys.argv[1] == "--trip-dir":
        if len(sys.argv) < 3:
            print("❌ 请提供行程目录路径")
            sys.exit(1)
        trip_dir = Path(sys.argv[2])
    else:
        trip_dir = Path(sys.argv[1])
    
    # 检查目录是否存在
    if not trip_dir.exists():
        print(f"❌ 目录不存在: {trip_dir}")
        sys.exit(1)
    
    if not trip_dir.is_dir():
        print(f"❌ 不是有效的目录: {trip_dir}")
        sys.exit(1)
    
    # 从目录名解析行程信息
    # 目录名格式: 出发地-目的地-MMDD-MMDD
    dir_name = trip_dir.name
    print(f"\n📁 行程目录: {dir_name}")
    
    # 尝试解析行程信息
    # 格式: 南京-北京-0208-0316
    parts = dir_name.split('-')
    if len(parts) >= 4:
        departure_city = parts[0]
        destination_city = parts[1]
        
        # 解析日期
        try:
            start_month = int(parts[2][:2])
            start_day = int(parts[2][2:])
            end_month = int(parts[3][:2])
            end_day = int(parts[3][2:])
            
            year = datetime.now().year
            start_date = datetime(year, start_month, start_day)
            end_date = datetime(year, end_month, end_day)
            
            # 创建行程段
            segment = TripSegment(
                departure_city=departure_city,
                destination_city=destination_city,
                start_date=start_date,
                end_date=end_date,
                transport_mode=None  # 从发票推断
            )
            
            segments = [segment]
            print(f"✅ 解析行程: {departure_city} → {destination_city}")
            print(f"   日期: {start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}")
            
        except (ValueError, IndexError) as e:
            print(f"⚠️ 无法从目录名解析行程信息: {e}")
            print("   将尝试从发票推断行程信息")
            segments = []
    else:
        print("⚠️ 目录名格式不符合预期，将尝试从发票推断行程信息")
        segments = []
    
    # 生成报销单
    result = generate_expense_report(trip_dir, segments)
    
    if result["success"]:
        print("\n✅ 报销单生成成功!")
        print(f"   Excel 文件: {result['excel_file']}")
        print(f"   匹配报告: {result['report_file']}")
        sys.exit(0)
    else:
        print(f"\n❌ 报销单生成失败: {result.get('error', '未知错误')}")
        sys.exit(1)


if __name__ == "__main__":
    main()