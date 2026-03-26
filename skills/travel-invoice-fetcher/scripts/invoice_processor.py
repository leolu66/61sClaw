#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
发票处理器（新流程）
实现用户要求的新发票处理流程
"""

import os
import re
import sys
import io
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from dataclasses import dataclass, field


# 设置 UTF-8 编码（Windows 兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


@dataclass
class InvoiceInfo:
    """发票信息"""
    file_path: Path
    invoice_no: Optional[str] = None
    amount: float = 0.0
    invoice_date: Optional[datetime] = None
    invoice_type: Optional[str] = None
    processed: bool = False


@dataclass
class TripInfo:
    """行程信息"""
    file_path: Path
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    start_time: Optional[datetime] = None
    city: Optional[str] = None
    start_point: Optional[str] = None
    end_point: Optional[str] = None
    amount: float = 0.0
    processed: bool = False


@dataclass
class RoadSegment:
    """路段对象"""
    segment_name: str
    segment_type: str  # 打车，高铁，地铁，大巴等
    invoice_info: InvoiceInfo = None
    trip_info: TripInfo = None
    # 其他补充信息
    departure_city: Optional[str] = None
    destination_city: Optional[str] = None
    
    def __str__(self):
        return f"{self.segment_name} ({self.segment_type})"


class InvoiceProcessor:
    """发票处理器（新流程）"""
    
    def __init__(self):
        self.pdf_files: List[Path] = []
        self.road_segments: List[RoadSegment] = []
        self.unprocessed_files: List[Path] = []
    
    def scan_pdf_files(self, trip_dir: Path) -> List[Path]:
        """扫描目录下所有PDF文件"""
        pdf_files = []
        for file_path in trip_dir.rglob('*.pdf'):
            if file_path.is_file():
                pdf_files.append(file_path)
                print(f"📄 找到PDF文件: {file_path.name}")
        return pdf_files
    
    def extract_amount_from_filename(self, filename: str) -> float:
        """从文件名提取金额"""
        patterns = [
            r'(\d+)_(\d+\.\d{2})_',   # 260315_34.00_北京滴滴
            r'_(\d+\.\d{2})_',       # _34.00_
        ]
        
        for pattern in patterns:
            match = re.search(pattern, filename)
            if match:
                try:
                    if pattern.startswith(r'(\d+)'):
                        amount_str = match.group(2)
                    else:
                        amount_str = match.group(1)
                    return float(amount_str)
                except (ValueError, IndexError):
                    continue
        
        return 0.0
    
    def extract_date_from_filename(self, filename: str) -> Optional[str]:
        """从文件名提取日期前缀"""
        match = re.search(r'^(\d{6})_', filename)
        if match:
            return match.group(1)
        return None
    
    def is_invoice_trip_pair(self, file1: Path, file2: Path) -> bool:
        """判断两个文件是否为发票和行程单的配对"""
        name1 = file1.name.lower()
        name2 = file2.name.lower()
        
        # 检查金额是否相同
        amount1 = self.extract_amount_from_filename(name1)
        amount2 = self.extract_amount_from_filename(name2)
        
        if abs(amount1 - amount2) > 0.01:
            return False
        
        # 检查日期前缀是否相同
        date1 = self.extract_date_from_filename(name1)
        date2 = self.extract_date_from_filename(name2)
        
        if date1 and date2 and date1 != date2:
            return False
        
        # 检查文件名是否包含发票和行程单的关键词
        has_invoice = any(keyword in name1 for keyword in ['发票', 'invoice']) or \
                      any(keyword in name2 for keyword in ['发票', 'invoice'])
        has_trip = any(keyword in name1 for keyword in ['行程单', 'trip', 'itinerary']) or \
                   any(keyword in name2 for keyword in ['行程单', 'trip', 'itinerary'])
        
        # 额外的判断：如果一个文件有金额而另一个没有，且金额不为0，则认为它们是配对的
        if amount1 > 0 and amount2 == 0:
            has_invoice = True
            has_trip = True
        elif amount1 == 0 and amount2 > 0:
            has_invoice = True
            has_trip = True
        
        return has_invoice and has_trip
    
    def round1_classification(self, pdf_files: List[Path]) -> List[RoadSegment]:
        """第一轮分类：根据文件名粗分类，归为一组"""
        print("\n🔍 第一轮分类：根据文件名分组发票和行程单")
        
        segments = []
        processed_indices = set()
        
        for i, file1 in enumerate(pdf_files):
            if i in processed_indices:
                continue
            
            for j, file2 in enumerate(pdf_files[i+1:], i+1):
                if j in processed_indices:
                    continue
                
                if self.is_invoice_trip_pair(file1, file2):
                    # 创建路段
                    amount = self.extract_amount_from_filename(file1.name)
                    
                    # 判断哪个是发票，哪个是行程单
                    name1 = file1.name.lower()
                    name2 = file2.name.lower()
                    
                    if any(keyword in name1 for keyword in ['发票', 'invoice']):
                        invoice_file = file1
                        trip_file = file2
                    else:
                        invoice_file = file2
                        trip_file = file1
                    
                    # 创建路段对象
                    segment = RoadSegment(
                        segment_name=f"{invoice_file.stem}",
                        segment_type="打车",  # 默认为打车，后续可调整
                        invoice_info=InvoiceInfo(file_path=invoice_file, amount=amount),
                        trip_info=TripInfo(file_path=trip_file, amount=amount)
                    )
                    
                    # 解析行程单信息
                    city, start_point, end_point, trip_date, start_time, trip_amount = self.parse_taxi_pdf(trip_file)
                    if segment.trip_info:
                        segment.trip_info.city = city
                        segment.trip_info.start_point = start_point
                        segment.trip_info.end_point = end_point
                        segment.trip_info.start_date = trip_date
                        segment.trip_info.start_time = start_time
                        if trip_amount > 0:
                            segment.trip_info.amount = trip_amount
                    
                    # 从发票文件提取金额
                    invoice_text = self.extract_text_from_pdf(invoice_file)
                    if invoice_text:
                        import re
                        # 从发票中提取金额
                        invoice_amount_patterns = [
                            r'总价[：:]\s*\D*\s*(\d+\.\d{2})',
                            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
                            r'[¥￥](\d+\.\d{2})',
                            r'合计(\d+\.\d{2})元',
                        ]
                        
                        for pattern in invoice_amount_patterns:
                            matches = re.findall(pattern, invoice_text)
                            if matches:
                                try:
                                    amounts = [float(m) for m in matches]
                                    invoice_amount = max(amounts)
                                    segment.invoice_info.amount = invoice_amount
                                    break
                                except ValueError:
                                    continue
                    
                    segments.append(segment)
                    processed_indices.add(i)
                    processed_indices.add(j)
                    
                    print(f"   ✅ 配对: {invoice_file.name} ↔ {trip_file.name}")
        
        # 收集未处理的文件
        self.unprocessed_files = [pdf_files[i] for i in range(len(pdf_files)) if i not in processed_indices]
        
        print(f"   已处理: {len(segments)} 组，未处理: {len(self.unprocessed_files)} 个文件")
        return segments
    
    def extract_text_from_pdf(self, file_path: Path) -> str:
        """从PDF文件提取文本"""
        try:
            import PyPDF2
            with open(file_path, 'rb') as f:
                reader = PyPDF2.PdfReader(f)
                text = ""
                for page in reader.pages:
                    text += page.extract_text() or ""
            return text
        except Exception as e:
            print(f"   ⚠️ 提取文本失败 {file_path.name}: {e}")
            return ""
    
    def is_train_invoice(self, file_path: Path) -> bool:
        """判断是否为高铁票"""
        filename = file_path.name.lower()
        
        # 首先通过文件名判断
        if any(keyword in filename for keyword in ['火车票', '高铁票', '铁路', '火车']):
            return True
        
        # 读取文件内容判断
        text = self.extract_text_from_pdf(file_path)
        if '铁路电子客票' in text or '中国铁路' in text:
            return True
        
        return False
    
    def is_taxi_invoice(self, file_path: Path) -> bool:
        """判断是否为打车票"""
        filename = file_path.name.lower()
        
        # 首先通过文件名判断
        if any(keyword in filename for keyword in ['滴滴', '出行', '打车', '久柏易游', '旅客运输', '行程单']):
            return True
        
        # 读取文件内容判断
        text = self.extract_text_from_pdf(file_path)
        if any(keyword in text for keyword in ['旅客运输服务', '滴滴', '出租车', '网约车', 'T3', '曹操']):
            return True
        
        return False
    
    def is_subway_invoice(self, file_path: Path) -> bool:
        """判断是否为地铁票"""
        filename = file_path.name.lower()
        
        # 首先通过文件名判断
        if any(keyword in filename for keyword in ['地铁', 'subway', 'metro']):
            return True
        
        # 读取文件内容判断
        text = self.extract_text_from_pdf(file_path)
        if any(keyword in text for keyword in ['地铁', '轨道交通', '地铁集团', '地铁公司']):
            return True
        
        return False
    
    def parse_train_pdf(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """高铁票PDF解析模板"""
        text = self.extract_text_from_pdf(file_path)
        
        # 提取路线
        departure = None
        destination = None
        
        # 模式1: 识别乱码的车站名称（如 ÄÏ¾©ÄÏ\nÕ¾ -> 南京南站）
        # 南京南站 -> ÄÏ¾©ÄÏ\nÕ¾
        # 北京南站 -> ±±¾©ÄÏ\nÕ¾
        if 'ÄÏ¾©ÄÏ' in text and '±±¾©ÄÏ' in text:
            # 确定哪个是出发站，哪个是到达站
            nanjing_pos = text.find('ÄÏ¾©ÄÏ')
            beijing_pos = text.find('±±¾©ÄÏ')
            
            if nanjing_pos < beijing_pos:
                departure = '南京南站'
                destination = '北京南站'
            else:
                departure = '北京南站'
                destination = '南京南站'
        else:
            # 正常中文模式
            route_patterns = [
                r'([\u4e00-\u9fa5]+?站)\s*[-→~到至]\s*([\u4e00-\u9fa5]+?站)',
                r'出发站[：:]\s*([\u4e00-\u9fa5]+?站).*?到达站[：:]\s*([\u4e00-\u9fa5]+?站)',
                r'([\u4e00-\u9fa5]+?)\s*[-→~到至]\s*([\u4e00-\u9fa5]+?)',
            ]
            
            for pattern in route_patterns:
                match = re.search(pattern, text)
                if match:
                    departure = match.group(1).replace('站', '')
                    destination = match.group(2).replace('站', '')
                    break
        
        # 提取日期
        trip_date = None
        
        # 模式1: 识别乱码日期（如 2026Äê02ÔÂ25ÈÕ -> 2026年02月25日）
        date_pattern = r'(\d{4})Äê(\d{1,2})ÔÂ(\d{1,2})ÈÕ'
        match = re.search(date_pattern, text)
        if match:
            try:
                year, month, day = map(int, match.groups())
                trip_date = datetime(year, month, day)
            except ValueError:
                pass
        
        # 模式2: 正常中文日期
        if not trip_date:
            date_patterns = [
                r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'(\d{4}年\d{1,2}月\d{1,2}日)',
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, text)
                if match:
                    try:
                        date_str = match.group(1)
                        if '年' in date_str:
                            trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
                        else:
                            trip_date = datetime.strptime(date_str, '%Y-%m-%d')
                        break
                    except ValueError:
                        continue
        
        # 提取时间（上车时间）
        start_time = None
        time_pattern = r'(\d{2}:\d{2})¿ª'  # 识别乱码时间（如 11:01¿ª -> 11:01开）
        match = re.search(time_pattern, text)
        if match:
            try:
                time_str = match.group(1)
                if trip_date:
                    # 合并日期和时间
                    hour, minute = map(int, time_str.split(':'))
                    start_time = datetime(trip_date.year, trip_date.month, trip_date.day, hour, minute)
            except ValueError:
                pass
        
        # 提取金额
        amount = 0.0
        amount_patterns = [
            r'价税合计[：:]\s*\D*\s*(\d+\.\d{2})',
            r'合计金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥]\s*(\d+\.\d{2})',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    amounts = [float(m) for m in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        return departure, destination, trip_date, start_time, amount
    
    def parse_taxi_pdf(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """打车票PDF解析模板"""
        filename = file_path.name.lower()
        
        # 根据文件名识别公司类型并应用相应的解析模板
        if '滴滴出行' in filename:
            return self._parse_didi_trip(file_path)
        elif '久柏易游' in filename:
            return self._parse_jiubai_trip(file_path)
        elif '吉利优行' in filename or '曹操出行' in filename:
            return self._parse_caocao_trip(file_path)
        else:
            # 默认解析模板
            return self._parse_generic_trip(file_path)
    
    def _parse_didi_trip(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """滴滴出行行程单解析模板"""
        text = self.extract_text_from_pdf(file_path)
        
        # 提取城市
        city = None
        city_patterns = [
            r'([\u4e00-\u9fa5]+?)滴滴出行',
            r'滴滴出行.*?([\u4e00-\u9fa5]+?)',
            r'城市\s*([\u4e00-\u9fa5]+?)',
        ]
        
        for pattern in city_patterns:
            match = re.search(pattern, text)
            if match:
                city = match.group(1)
                break
        
        # 提取出发日期（从行程起止日期中提取）
        trip_date = None
        date_pattern = r'行程起止日期：(\d{4}-\d{2}-\d{2})\s*至'
        match = re.search(date_pattern, text)
        if match:
            try:
                date_str = match.group(1)
                trip_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                pass
        
        # 如果没有找到行程起止日期，尝试其他日期模式
        if not trip_date:
            date_patterns = [
                r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'(\d{4}年\d{1,2}月\d{1,2}日)',
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, text)
                if match:
                    try:
                        date_str = match.group(1)
                        if '年' in date_str:
                            trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
                        else:
                            trip_date = datetime.strptime(date_str, '%Y-%m-%d')
                        break
                    except ValueError:
                        continue
        
        # 提取上车时间（从行程明细表格中提取）
        start_time = None
        time_pattern = r'\d+\s+[^\s]+\s+(\d{2}-\d{2}\s+\d{2}:\d{2})'
        match = re.search(time_pattern, text)
        if match and trip_date:
            try:
                time_str = match.group(1)
                # 解析格式：02-25 19:54
                date_part, time_part = time_str.split()
                month, day = map(int, date_part.split('-'))
                hour, minute = map(int, time_part.split(':'))
                start_time = datetime(trip_date.year, month, day, hour, minute)
            except ValueError:
                pass
        
        # 如果没有找到表格中的上车时间，尝试其他时间模式
        if not start_time:
            time_patterns = [
                r'上车时间[：:]\s*(\d{2}:\d{2})',
                r'(\d{2}:\d{2})\s*上车',
            ]
            
            for pattern in time_patterns:
                match = re.search(pattern, text)
                if match and trip_date:
                    try:
                        time_str = match.group(1)
                        hour, minute = map(int, time_str.split(':'))
                        start_time = datetime(trip_date.year, trip_date.month, trip_date.day, hour, minute)
                    except ValueError:
                        continue
        
        # 提取起点和终点（从行程明细表格中提取）
        start_point = None
        end_point = None
        # 表格格式：序号 车型 上车时间 星期 城市 起点 终点 里程 金额
        # 示例：1 特惠快车 07-31 14:20 周三 北京市 一品·亦庄-南门 丽泽桥|丽泽天地购物中心 23.4 39.77
        route_pattern = r'\d+\s+[^\s]+\s+\d{2}-\d{2}\s+\d{2}:\d{2}\s+[^\s]+\s+[^\s]+\s+(.+?)\s+(.+?)\s+\d+\.\d+'
        match = re.search(route_pattern, text)
        if match:
            start_point = match.group(1)
            end_point = match.group(2)
        
        # 如果没有找到表格中的路线，尝试其他模式
        if not start_point or not end_point:
            route_patterns = [
                r'起点[：:]\s*([^\n]+?)\s*终点',
                r'上车[：:]\s*([^\n]+?)\s*下车',
            ]
            
            for pattern in route_patterns:
                match = re.search(pattern, text)
                if match:
                    route_parts = match.group(1).split('到')
                    if len(route_parts) >= 2:
                        start_point = route_parts[0].strip()
                        end_point = route_parts[1].strip()
                    break
        
        # 提取金额
        amount = 0.0
        amount_patterns = [
            r'总价[：:]\s*\D*\s*(\d+\.\d{2})',
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥](\d+\.\d{2})',
            r'合计(\d+\.\d{2})元',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    amounts = [float(m) for m in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        return city, start_point, end_point, trip_date, start_time, amount
    
    def _parse_jiubai_trip(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """久柏易游及第三方打车公司行程单解析模板"""
        text = self.extract_text_from_pdf(file_path)
        
        # 提取城市
        city = None
        city_patterns = [
            r'([\u4e00-\u9fa5]+?)久柏易游',
            r'久柏易游.*?([\u4e00-\u9fa5]+?)',
            r'城市\s*([\u4e00-\u9fa5]+?)',
        ]
        
        for pattern in city_patterns:
            match = re.search(pattern, text)
            if match:
                city = match.group(1)
                break
        
        # 提取出发日期（从行程起止日期中提取）
        trip_date = None
        date_pattern = r'行程起止日期：(\d{4}-\d{2}-\d{2})\s*至'
        match = re.search(date_pattern, text)
        if match:
            try:
                date_str = match.group(1)
                trip_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                pass
        
        # 如果没有找到行程起止日期，尝试其他日期模式
        if not trip_date:
            date_patterns = [
                r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'(\d{4}年\d{1,2}月\d{1,2}日)',
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, text)
                if match:
                    try:
                        date_str = match.group(1)
                        if '年' in date_str:
                            trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
                        else:
                            trip_date = datetime.strptime(date_str, '%Y-%m-%d')
                        break
                    except ValueError:
                        continue
        
        # 提取上车时间（从行程明细表格中提取）
        start_time = None
        time_pattern = r'\d+\s+[^\s]+\s+(\d{2}-\d{2}\s+\d{2}:\d{2})\s*[^\n]*'
        match = re.search(time_pattern, text)
        if match and trip_date:
            try:
                time_str = match.group(1)
                # 解析格式：02-25 09:25
                date_part, time_part = time_str.split()
                month, day = map(int, date_part.split('-'))
                hour, minute = map(int, time_part.split(':'))
                start_time = datetime(trip_date.year, month, day, hour, minute)
            except ValueError:
                pass
        
        # 如果没有找到表格中的上车时间，尝试其他时间模式
        if not start_time:
            time_patterns = [
                r'上车时间[：:]\s*(\d{2}:\d{2})',
                r'(\d{2}:\d{2})\s*上车',
            ]
            
            for pattern in time_patterns:
                match = re.search(pattern, text)
                if match and trip_date:
                    try:
                        time_str = match.group(1)
                        hour, minute = map(int, time_str.split(':'))
                        start_time = datetime(trip_date.year, trip_date.month, trip_date.day, hour, minute)
                    except ValueError:
                        continue
        
        # 提取起点和终点（从行程明细表格中提取）
        start_point = None
        end_point = None
        # 表格格式：序号 车型 上车时间 星期 城市 起点 终点 里程 金额
        route_pattern = r'\d+\s+[^\s]+\s+\d{2}-\d{2}\s+\d{2}:\d{2}\s+[^\s]+\s+[^\s]+\s+(.+?)\s+(.+?)\s+\d+\.\d+'
        match = re.search(route_pattern, text)
        if match:
            start_point = match.group(1)
            end_point = match.group(2)
        
        # 如果没有找到表格中的路线，尝试其他模式
        if not start_point or not end_point:
            route_patterns = [
                r'起点\s*[:：]\s*([^\n]+?)\s*终点',
                r'上车\s*[:：]\s*([^\n]+?)\s*下车',
            ]
            
            for pattern in route_patterns:
                match = re.search(pattern, text)
                if match:
                    route_parts = match.group(1).split('到')
                    if len(route_parts) >= 2:
                        start_point = route_parts[0].strip()
                        end_point = route_parts[1].strip()
                    break
        
        # 提取金额
        amount = 0.0
        amount_patterns = [
            r'总价[：:]\s*\D*\s*(\d+\.\d{2})',
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥](\d+\.\d{2})',
            r'合计(\d+\.\d{2})元',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    amounts = [float(m) for m in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        return city, start_point, end_point, trip_date, start_time, amount
    
    def _parse_caocao_trip(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """曹操出行行程单解析模板"""
        text = self.extract_text_from_pdf(file_path)
        
        # 提取城市
        city = None
        
        # 提取日期
        trip_date = None
        date_patterns = [
            r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
            r'(\d{4}年\d{1,2}月\d{1,2}日)',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    date_str = match.group(1)
                    if '年' in date_str:
                        trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
                    else:
                        trip_date = datetime.strptime(date_str, '%Y-%m-%d')
                    break
                except ValueError:
                    continue
        
        # 提取时间
        start_time = None
        
        # 提取路线
        start_point = None
        end_point = None
        
        # 提取金额
        amount = 0.0
        amount_patterns = [
            r'总价[：:]\s*\D*\s*(\d+\.\d{2})',
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥]\s*(\d+\.\d{2})',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    amounts = [float(m) for m in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        return city, start_point, end_point, trip_date, start_time, amount
    
    def _parse_generic_trip(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """通用行程单解析模板"""
        text = self.extract_text_from_pdf(file_path)
        
        # 提取城市
        city = None
        
        # 检查是否为地铁出行行程单
        is_subway = '地铁出行' in text
        
        # 提取日期
        trip_date = None
        
        # 首先尝试从行程明细表格中提取日期（地铁行程单格式）
        date_pattern = r'\d+\s+(\d{4}年\d{1,2}月\d{1,2}日)\s+'
        match = re.search(date_pattern, text)
        if match:
            try:
                date_str = match.group(1)
                trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
            except ValueError:
                pass
        
        # 如果没有找到表格中的日期，尝试其他日期模式
        if not trip_date:
            date_patterns = [
                r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
                r'(\d{4}年\d{1,2}月\d{1,2}日)',
                r'申请日期[：:]\s*(\d{4}年\d{1,2}月\d{1,2}日)',
            ]
            
            for pattern in date_patterns:
                match = re.search(pattern, text)
                if match:
                    try:
                        date_str = match.group(1)
                        if '年' in date_str:
                            trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
                        else:
                            trip_date = datetime.strptime(date_str, '%Y-%m-%d')
                        break
                    except ValueError:
                        continue
        
        # 提取上车时间
        start_time = None
        
        # 尝试从行程明细表格中提取时间（地铁行程单格式）
        time_pattern = r'\d+\s+\d{4}年\d{1,2}月\d{1,2}日\s+(\d{2}:\d{2})-'
        match = re.search(time_pattern, text)
        if match and trip_date:
            try:
                time_str = match.group(1)
                hour, minute = map(int, time_str.split(':'))
                start_time = datetime(trip_date.year, trip_date.month, trip_date.day, hour, minute)
            except ValueError:
                pass
        
        # 如果没有找到表格中的时间，尝试其他时间模式
        if not start_time:
            time_patterns = [
                r'上车时间[：:]\s*(\d{2}:\d{2})',
                r'(\d{2}:\d{2})\s*上车',
            ]
            
            for pattern in time_patterns:
                match = re.search(pattern, text)
                if match and trip_date:
                    try:
                        time_str = match.group(1)
                        hour, minute = map(int, time_str.split(':'))
                        start_time = datetime(trip_date.year, trip_date.month, trip_date.day, hour, minute)
                    except ValueError:
                        continue
        
        # 提取路线（起点和终点）
        start_point = None
        end_point = None
        
        # 尝试从行程明细表格中提取行程站点（地铁行程单格式）
        route_pattern = r'\d+\s+\d{4}年\d{1,2}月\d{1,2}日\s+\d{2}:\d{2}-\d{2}:\d{2}\s+([^ ]+?-[^ ]+?)\s+'
        match = re.search(route_pattern, text)
        if match:
            route = match.group(1)
            # 分割起点和终点
            if '-' in route:
                start_point, end_point = route.split('-', 1)
        
        # 如果没有找到表格中的路线，尝试其他模式
        if not start_point or not end_point:
            route_patterns = [
                r'起点[：:]\s*([^\n]+?)\s*终点',
                r'上车[：:]\s*([^\n]+?)\s*下车',
                r'行程站点[：:]\s*([^\n]+?)\s*金额',
            ]
            
            for pattern in route_patterns:
                match = re.search(pattern, text)
                if match:
                    route = match.group(1)
                    if '-' in route:
                        start_point, end_point = route.split('-', 1)
                    break
        
        # 提取金额
        amount = 0.0
        amount_patterns = [
            r'总价[：:]\s*\D*\s*(\d+\.\d{2})',
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥](\d+\.\d{2})',
            r'合计(\d+\.\d{2})元',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    amounts = [float(m) for m in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        return city, start_point, end_point, trip_date, start_time, amount
    
    def parse_subway_pdf(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """地铁票PDF解析模板"""
        text = self.extract_text_from_pdf(file_path)
        
        # 提取城市
        city = None
        city_patterns = [
            r'([\u4e00-\u9fa5]+?)地铁集团',
            r'地铁集团.*?([\u4e00-\u9fa5]+?)',
        ]
        
        for pattern in city_patterns:
            match = re.search(pattern, text)
            if match:
                city = match.group(1)
                break
        
        # 提取日期
        trip_date = None
        date_patterns = [
            r'(\d{4}[-/]\d{1,2}[-/]\d{1,2})',
            r'(\d{4}年\d{1,2}月\d{1,2}日)',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    date_str = match.group(1)
                    if '年' in date_str:
                        trip_date = datetime.strptime(date_str, '%Y年%m月%d日')
                    else:
                        trip_date = datetime.strptime(date_str, '%Y-%m-%d')
                    break
                except ValueError:
                    continue
        
        # 提取进站时间
        start_time = None
        time_patterns = [
            r'进站时间[：:]\s*(\d{2}:\d{2})',
            r'(\d{2}:\d{2})\s*进站',
        ]
        
        for pattern in time_patterns:
            match = re.search(pattern, text)
            if match:
                try:
                    time_str = match.group(1)
                    if trip_date:
                        hour, minute = map(int, time_str.split(':'))
                        start_time = datetime(trip_date.year, trip_date.month, trip_date.day, hour, minute)
                except ValueError:
                    continue
        
        # 提取起点和终点（地铁站名）
        start_point = None
        end_point = None
        route_patterns = [
            r'起点[：:]\s*([^\n]+?)\s*终点',
            r'进站[：:]\s*([^\n]+?)\s*出站',
        ]
        
        for pattern in route_patterns:
            match = re.search(pattern, text)
            if match:
                route_parts = match.group(1).split('到')
                if len(route_parts) >= 2:
                    start_point = route_parts[0].strip()
                    end_point = route_parts[1].strip()
                break
        
        # 提取金额
        amount = 0.0
        amount_patterns = [
            r'金额[：:]\s*\D*\s*(\d+\.\d{2})',
            r'票价[：:]\s*\D*\s*(\d+\.\d{2})',
            r'[¥￥]\s*(\d+\.\d{2})',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text)
            if matches:
                try:
                    amounts = [float(m) for m in matches]
                    amount = max(amounts)
                    break
                except ValueError:
                    continue
        
        return city, start_point, end_point, trip_date, start_time, amount
    
    def extract_train_info(self, file_path: Path) -> Tuple[Optional[str], Optional[str], Optional[datetime], Optional[datetime], float]:
        """从高铁票提取信息"""
        # 调用高铁票解析模板
        return self.parse_train_pdf(file_path)
    
    def round2_classification(self) -> List[RoadSegment]:
        """第二轮分类：识别高铁路段"""
        print("\n🚄 第二轮分类：识别高铁路段")
        
        segments = []
        processed_indices = set()
        
        for i, file_path in enumerate(self.unprocessed_files):
            if self.is_train_invoice(file_path):
                # 提取高铁信息
                departure, destination, trip_date, start_time, amount = self.extract_train_info(file_path)
                
                segment = RoadSegment(
                    segment_name=f"{departure or '未知'}到{destination or '未知'}",
                    segment_type="高铁",
                    invoice_info=InvoiceInfo(
                        file_path=file_path,
                        amount=amount,
                        invoice_date=trip_date,
                        invoice_type="高铁票"
                    ),
                    trip_info=TripInfo(
                        file_path=file_path,
                        start_date=trip_date,
                        end_date=trip_date,
                        start_time=start_time,  # 添加上车时间
                        start_point=departure,
                        end_point=destination,
                        amount=amount
                    ),
                    departure_city=departure,
                    destination_city=destination
                )
                
                segments.append(segment)
                processed_indices.add(i)
                print(f"   ✅ 识别高铁票: {file_path.name}")
        
        # 更新未处理文件列表
        self.unprocessed_files = [self.unprocessed_files[i] for i in range(len(self.unprocessed_files)) if i not in processed_indices]
        
        print(f"   已识别: {len(segments)} 张高铁票，剩余未处理: {len(self.unprocessed_files)} 个文件")
        return segments
    
    def get_file_mtime(self, file_path: Path) -> float:
        """获取文件修改时间"""
        try:
            return file_path.stat().st_mtime
        except Exception:
            return 0.0
    
    def round3_classification(self) -> List[RoadSegment]:
        """第三轮分类：通过金额相同和时间相近匹配发票和行程"""
        print("\n🔗 第三轮分类：匹配剩余发票和行程")
        
        segments = []
        processed_indices = set()
        
        # 将文件分为可能的发票和可能的行程单
        potential_invoices = []
        potential_trips = []
        
        for file_path in self.unprocessed_files:
            filename = file_path.name.lower()
            if any(keyword in filename for keyword in ['发票', 'invoice']):
                potential_invoices.append(file_path)
            elif any(keyword in filename for keyword in ['行程单', 'trip', 'itinerary']):
                potential_trips.append(file_path)
            else:
                # 无法确定，暂放在发票列表中
                potential_invoices.append(file_path)
        
        # 按金额和时间匹配
        for invoice_file in potential_invoices:
            invoice_amount = self.extract_amount_from_filename(invoice_file.name)
            invoice_time = self.get_file_mtime(invoice_file)
            
            for trip_file in potential_trips:
                trip_amount = self.extract_amount_from_filename(trip_file.name)
                trip_time = self.get_file_mtime(trip_file)
                
                # 金额相同（允许0.01元误差）且时间相近（5分钟内）
                if abs(invoice_amount - trip_amount) < 0.01 and abs(invoice_time - trip_time) < 300:
                    segment = RoadSegment(
                        segment_name=f"{invoice_file.stem}",
                        segment_type="打车",  # 默认打车，后续可调整
                        invoice_info=InvoiceInfo(
                            file_path=invoice_file,
                            amount=invoice_amount
                        ),
                        trip_info=TripInfo(
                            file_path=trip_file,
                            amount=trip_amount
                        )
                    )
                    
                    # 解析行程单信息
                    city, start_point, end_point, trip_date, start_time, trip_amount = self.parse_taxi_pdf(trip_file)
                    if segment.trip_info:
                        segment.trip_info.city = city
                        segment.trip_info.start_point = start_point
                        segment.trip_info.end_point = end_point
                        segment.trip_info.start_date = trip_date
                        segment.trip_info.start_time = start_time
                    
                    segments.append(segment)
                    print(f"   ✅ 匹配: {invoice_file.name} ↔ {trip_file.name}")
                    break
        
        print(f"   已匹配: {len(segments)} 组")
        return segments
    
    def sort_segments(self, segments: List[RoadSegment]) -> List[RoadSegment]:
        """排序路段：先按出发城市和目的地城市分两段，中间是高铁或飞机，每段按日期和时间排序"""
        print("\n📊 路段排序")
        
        # 识别主要的出发城市和目的地城市
        departure_cities = []
        destination_cities = []
        
        for segment in segments:
            if segment.segment_type == "高铁":
                if segment.departure_city:
                    departure_cities.append(segment.departure_city)
                if segment.destination_city:
                    destination_cities.append(segment.destination_city)
        
        # 统计出现频率最高的城市
        city_counts = {}
        for city in departure_cities + destination_cities:
            if city:
                city_counts[city] = city_counts.get(city, 0) + 1
        
        # 确定主要城市
        main_departure = None
        main_destination = None
        if city_counts:
            # 假设第一个高铁路段的出发地是主要出发城市
            for segment in segments:
                if segment.segment_type == "高铁" and segment.departure_city:
                    main_departure = segment.departure_city
                    main_destination = segment.destination_city
                    break
        
        # 排序函数
        def sort_key(segment):
            # 第一优先级：高铁优先排在中间
            if segment.segment_type == "高铁":
                priority = 2
            # 第二优先级：出发城市段（打车）
            elif main_departure and segment.trip_info and main_departure in str(segment.trip_info.file_path):
                priority = 1
            # 第三优先级：目的地城市段（打车）
            elif main_destination and segment.trip_info and main_destination in str(segment.trip_info.file_path):
                priority = 3
            else:
                priority = 4
            
            # 第二排序键：日期和时间
            date_key = segment.trip_info.start_date if segment.trip_info and segment.trip_info.start_date else datetime(1900, 1, 1)
            time_key = segment.trip_info.start_time if segment.trip_info and segment.trip_info.start_time else datetime(1900, 1, 1)
            
            return (priority, date_key, time_key)
        
        sorted_segments = sorted(segments, key=sort_key)
        
        # 打印排序结果
        print("   排序结果:")
        for i, segment in enumerate(sorted_segments, 1):
            print(f"   {i}. {segment}")
        
        return sorted_segments
    
    def generate_expense_report(self, trip_dir: Path, segments: List[RoadSegment]) -> Path:
        """生成报销单Excel"""
        print("\n📑 生成报销单")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            print("   ⚠️ 未安装 openpyxl，尝试安装...")
            os.system(f"{sys.executable} -m pip install openpyxl -q")
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "差旅报销单"
        
        # 设置标题
        ws['A1'] = "差旅报销单"
        ws.merge_cells('A1:F1')
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置表头
        headers = ["序号", "路线", "出发日期", "上车时间", "交通工具", "金额(元)", "发票文件"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 填充数据
        for row, segment in enumerate(segments, 4):
            ws.cell(row=row, column=1, value=row - 3)
            
            # 路线
            route = segment.segment_name
            if segment.trip_info and segment.trip_info.start_point and segment.trip_info.end_point:
                route = f"{segment.trip_info.start_point} → {segment.trip_info.end_point}"
            ws.cell(row=row, column=2, value=route)
            
            # 出发日期
            departure_date = ""
            if segment.trip_info and segment.trip_info.start_date:
                departure_date = segment.trip_info.start_date.strftime('%Y-%m-%d')
            ws.cell(row=row, column=3, value=departure_date)
            
            # 上车时间
            start_time = ""
            if segment.trip_info and segment.trip_info.start_time:
                start_time = segment.trip_info.start_time.strftime('%H:%M')
            ws.cell(row=row, column=4, value=start_time)
            
            # 交通工具
            ws.cell(row=row, column=5, value=segment.segment_type)
            
            # 金额
            amount = segment.invoice_info.amount if segment.invoice_info else 0.0
            ws.cell(row=row, column=6, value=amount)
            ws.cell(row=row, column=6).number_format = '0.00'
            
            # 发票文件
            invoice_file = segment.invoice_info.file_path.name if segment.invoice_info else ""
            ws.cell(row=row, column=7, value=invoice_file)
        
        # 添加合计行
        total_row = len(segments) + 4
        ws.cell(row=total_row, column=1, value="合计")
        ws.merge_cells(f'A{total_row}:E{total_row}')
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        
        total_amount = sum(seg.invoice_info.amount if seg.invoice_info else 0.0 for seg in segments)
        ws.cell(row=total_row, column=6, value=total_amount)
        ws.cell(row=total_row, column=6).number_format = '0.00'
        ws.cell(row=total_row, column=6).font = Font(bold=True)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 40
        
        # 添加边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(3, total_row + 1):
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = thin_border
        
        # 保存文件
        timestamp = datetime.now().strftime('%H%M%S')
        output_file = trip_dir / f"报销单_新流程_{timestamp}.xlsx"
        wb.save(str(output_file))
        
        print(f"   ✅ 报销单已生成: {output_file}")
        return output_file
    
    def process(self, trip_dir: Path) -> Path:
        """执行完整的发票处理流程"""
        print("🚀 开始发票处理流程")
        print("=" * 50)
        
        # 1. 扫描PDF文件
        pdf_files = self.scan_pdf_files(trip_dir)
        if not pdf_files:
            print("❌ 未找到PDF文件")
            return None
        
        # 2. 第一轮分类
        segments_round1 = self.round1_classification(pdf_files)
        
        # 3. 第二轮分类
        segments_round2 = self.round2_classification()
        
        # 4. 第三轮分类
        segments_round3 = self.round3_classification()
        
        # 合并所有路段
        all_segments = segments_round1 + segments_round2 + segments_round3
        
        if not all_segments:
            print("❌ 未识别到任何路段")
            return None
        
        # 5. 排序路段
        sorted_segments = self.sort_segments(all_segments)
        
        # 6. 生成报销单
        report_file = self.generate_expense_report(trip_dir, sorted_segments)
        
        print("\n" + "=" * 50)
        print("✅ 发票处理流程完成")
        print("=" * 50)
        
        return report_file


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print_help()
        sys.exit(1)
    
    arg1 = sys.argv[1]
    if arg1 in ['--help', '-h', 'help']:
        print_help()
        sys.exit(0)
    
    trip_dir = Path(arg1)
    if not trip_dir.exists():
        print(f"❌ 目录不存在: {trip_dir}")
        sys.exit(1)
    
    processor = InvoiceProcessor()
    processor.process(trip_dir)


def print_help():
    """打印帮助信息"""
    print("发票处理器（新流程）")
    print("=" * 50)
    print("使用方法:")
    print(f"  python {sys.argv[0]} <行程目录路径>")
    print(f"  python {sys.argv[0]} --help | -h")
    print("\n示例:")
    print(f"  python {sys.argv[0]} 'D:\\Users\\luzhe\\报销凭证\\南京-北京-0208-0316'")
    print("\n功能说明:")
    print("  1. 扫描目录下所有PDF文件")
    print("  2. 第一轮分类：根据文件名分组发票和行程单")
    print("  3. 第二轮分类：识别高铁路段")
    print("  4. 第三轮分类：通过金额和时间匹配剩余发票")
    print("  5. 排序路段并生成报销单")


if __name__ == "__main__":
    main()
