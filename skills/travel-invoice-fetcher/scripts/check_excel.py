#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查生成的Excel文件中的数据
"""

import sys
import io
from pathlib import Path

# 设置 UTF-8 编码（Windows 兼容）
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')


def check_excel_file():
    """检查Excel文件中的数据"""
    
    # 设置测试目录
    trip_dir = Path(r'D:\Users\luzhe\报销凭证\南京-北京-0225-0311')
    
    if not trip_dir.exists():
        print(f"❌ 目录不存在: {trip_dir}")
        return
    
    # 查找最新的报销单文件
    excel_files = []
    for file_path in trip_dir.glob('报销单_新流程_*.xlsx'):
        excel_files.append(file_path)
    
    if not excel_files:
        print("❌ 未找到报销单文件")
        return
    
    # 按修改时间排序，获取最新的文件
    excel_files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    latest_file = excel_files[0]
    
    print(f"检查文件: {latest_file.name}")
    print("=" * 70)
    
    try:
        import openpyxl
        
        # 加载Excel文件
        wb = openpyxl.load_workbook(latest_file)
        ws = wb.active
        
        # 获取表头
        headers = []
        for cell in ws[3]:
            headers.append(cell.value)
        
        print("表头:", headers)
        
        # 获取数据行
        print("\n数据内容:")
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row[0] is None:
                continue
            print(f"  序号: {row[0]}, 路线: {row[1]}, 出发日期: {row[2]}, 上车时间: {row[3]}, 交通工具: {row[4]}, 金额(元): {row[5]}, 发票文件: {row[6]}")
    
    except Exception as e:
        print(f"❌ 读取Excel文件失败: {e}")


if __name__ == "__main__":
    check_excel_file()
