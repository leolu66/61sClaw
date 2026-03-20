# -*- coding: utf-8 -*-
"""
生成报销单Excel
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "报销单"

# 样式定义
header_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
sub_header_font = Font(name='微软雅黑', size=11, bold=True)
sub_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
normal_font = Font(name='微软雅黑', size=10)
border_thin = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ===== 第1行：标题 =====
ws.merge_cells('A1:J1')
ws['A1'] = "报销单 - 南京→北京 (2026-02-25 至 2026-03-11)"
ws['A1'].font = header_font
ws['A1'].alignment = center_align
ws.row_dimensions[1].height = 30

# ===== 第2行：表头 =====
ws['A2'] = "序号"
ws['B2'] = "行程"
ws['C2'] = "阶段"
ws['D2'] = "出发地"
ws['E2'] = "到达地"
ws['F2'] = "交通方式"
ws['G2'] = "金额(元)"
ws['H2'] = "发票号"
ws['I2'] = "发票文件"
ws['J2'] = "备注"

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    cell = ws[f'{col}2']
    cell.font = sub_header_font
    cell.fill = sub_header_fill
    cell.alignment = center_align
    cell.border = border_thin

# ===== 数据行 =====
# 往程
trip_data = [
    # 序号, 行程, 阶段, 出发地, 到达地, 交通方式, 金额, 发票号, 发票文件, 备注
    (1, '往', '1.家→中转', '家', '地铁站', '打车', 24.39, '', '260315_34.00_滴滴行程单_明细.pdf', '去地铁站'),
    (2, '往', '2.中转→站', '地铁站', '南京南站', '地铁', 0, '', '', ''),
    (3, '往', '3.长途交通', '南京南站', '北京南站', '高铁', 533.0, '26329116804002454260', '火车票.pdf', '公司统一报销'),
    (4, '往', '4.到→中转', '北京南站', '地铁站', '地铁', 0, '', '', ''),
    (5, '往', '5.中转→宿', '地铁站', '宿舍', '打车', 28.43, '', '260316_28.43_滴滴行程单_明细.pdf', ''),
    # 返程（根据实际行程，如果有的话）
    (6, '返', '1.宿→中转', '宿舍', '地铁站', '打车', 0, '', '', ''),
    (7, '返', '2.中转→站', '地铁站', '北京南站', '地铁', 0, '', '', ''),
    (8, '返', '3.长途交通', '北京南站', '南京南站', '高铁', 0, '', '', '公司统一报销'),
    (9, '返', '4.到→中转', '南京南站', '地铁站', '地铁', 0, '', '', ''),
    (10, '返', '5.中转→家', '地铁站', '家', '打车', 13.0, '', '260315_34.00_滴滴行程单.pdf', '预估'),
]

row = 3
for data in trip_data:
    ws[f'A{row}'] = data[0]  # 序号
    ws[f'B{row}'] = data[1]  # 行程
    ws[f'C{row}'] = data[2]  # 阶段
    ws[f'D{row}'] = data[3]  # 出发地
    ws[f'E{row}'] = data[4]  # 到达地
    ws[f'F{row}'] = data[5]  # 交通方式
    ws[f'G{row}'] = data[6]  # 金额
    ws[f'H{row}'] = data[7]  # 发票号
    ws[f'I{row}'] = data[8]  # 发票文件
    ws[f'J{row}'] = data[9]  # 备注
    
    # 格式化
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        cell = ws[f'{col}{row}']
        cell.font = normal_font
        cell.border = border_thin
        if col in ['A', 'B', 'C', 'F']:
            cell.alignment = center_align
        elif col in ['G']:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = left_align
    
    row += 1

# ===== 合计行 =====
ws.merge_cells(f'A{row}:F{row}')
ws[f'A{row}'] = "合计"
ws[f'A{row}'].font = Font(name='微软雅黑', size=11, bold=True)
ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')

ws[f'G{row}'] = f"=SUM(G3:G{row-1})"
ws[f'G{row}'].font = Font(name='微软雅黑', size=11, bold=True)
ws[f'G{row}'].number_format = '0.00'
ws[f'G{row}'].alignment = Alignment(horizontal='right', vertical='center')

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws[f'{col}{row}'].border = border_thin
    ws[f'{col}{row}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

# ===== 列宽设置 =====
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 20
ws.column_dimensions['I'].width = 25
ws.column_dimensions['J'].width = 15

# ===== 保存 =====
output_path = r'D:\Users\luzhe\报销凭证\南京-北京-0225-0311\报销单.xlsx'
wb.save(output_path)
print(f"报销单已生成: {output_path}")
