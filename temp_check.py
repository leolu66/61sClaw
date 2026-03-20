# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook(r'D:\Users\luzhe\报销凭证\南京-北京-0225-0311\报销单_自动.xlsx')
ws = wb.active

print('Excel auto-generated:')
for row in range(1, 16):
    values = []
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        v = ws[f'{col}{row}'].value
        if v is None:
            v = ''
        values.append(str(v))
    print(f'Row {row}: A={values[0]} B={values[1]} C={values[2]} G={values[6]} I={values[8]}')
