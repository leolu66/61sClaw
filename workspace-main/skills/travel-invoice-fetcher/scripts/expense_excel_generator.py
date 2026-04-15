# -*- coding: utf-8 -*-
"""
报销单Excel自动生成模块
功能：根据发票目录自动生成报销单，自动匹配发票
"""

import os
import json
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 导入发票归集模块
from invoice_aggregator import InvoiceAggregator


class ExpenseExcelGenerator:
    """报销单Excel生成器"""
    
    def __init__(self, trip_dir: str):
        """
        初始化
        
        Args:
            trip_dir: 行程目录路径
        """
        self.trip_dir = Path(trip_dir)
        self.trip_name = self.trip_dir.name
        
        # 行程信息（从目录名解析）
        self.departure = ''
        self.destination = ''
        self.start_date = ''
        self.end_date = ''
        self._parse_trip_name()
        
        # 发票列表
        self.invoices = []  # 发票
        self.trips = []     # 行程单
        self._load_invoices()
    
    def _parse_trip_name(self):
        """从目录名解析行程信息"""
        # 格式: 南京-北京-0225-0311
        name = self.trip_name
        parts = name.split('-')
        if len(parts) >= 4:
            self.departure = parts[0]
            self.destination = parts[1]
            self.start_date = parts[2]
            self.end_date = parts[3]
    
    def _load_invoices(self):
        """加载发票数据"""
        # 读取 invoice_summary.json（如果存在）
        summary_file = self.trip_dir / 'invoice_summary.json'
        
        if summary_file.exists():
            with open(summary_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            for inv in data.get('invoices', []):
                if inv['source_format'] == 'pdf':
                    filename = inv.get('source_file', '')
                    # 排除行程单（文件名含"行程单"）
                    if '行程单' in filename:
                        self.trips.append(inv)
                    else:
                        self.invoices.append(inv)
        else:
            # 如果没有汇总文件，直接扫描
            aggregator = InvoiceAggregator(str(self.trip_dir))
            aggregator.scan_directory()
            aggregator.deduplicate()
            
            for inv in aggregator.invoices:
                if inv.source_format == 'pdf':
                    filename = inv.source_file
                    # 排除行程单
                    if '行程单' in filename:
                        self.trips.append(inv.to_dict())
                    else:
                        self.invoices.append(inv.to_dict())
    
    def _match_invoice(self, expected_amount: float, invoice_type: str = 'taxi') -> Optional[Dict]:
        """
        根据期望金额匹配发票
        
        Args:
            expected_amount: 期望金额
            invoice_type: 发票类型
            
        Returns:
            匹配的发票字典，没有则返回None
        """
        if not self.invoices:
            return None
        
        # 筛选同类型发票
        candidates = [inv for inv in self.invoices if inv['invoice_type'] == invoice_type]
        
        if not candidates:
            candidates = self.invoices  # 如果没有同类型，用所有发票
        
        # 找金额最接近的
        best_match = None
        min_diff = float('inf')
        
        for inv in candidates:
            amount = inv['amount']
            if amount > 0:  # 跳过金额为0的
                diff = abs(amount - expected_amount)
                if diff < min_diff:
                    min_diff = diff
                    best_match = inv
        
        return best_match
    
    def _get_trip_remark(self, matched_invoice: Dict) -> str:
        """
        获取行程单备注
        
        Args:
            matched_invoice: 匹配到的发票
            
        Returns:
            行程单备注字符串
        """
        if not self.trips:
            return ''
        
        invoice_file = matched_invoice.get('source_file', '')
        
        # 找文件名相似的行程单
        # 例如：发票是 260315_34.00_xxx.pdf，行程单是 260315_34.00_xxx_行程单.pdf
        base_name = invoice_file.replace('.pdf', '')
        
        for trip in self.trips:
            trip_file = trip.get('source_file', '')
            # 检查是否是同一笔（通过日期+金额判断）
            if trip.get('date') == matched_invoice.get('date'):
                # 行程单作为备注
                return f"行程单: {trip_file}"
        
        return ''
    
    def _load_trip_info(self) -> Dict:
        """加载行程信息"""
        trip_info_file = self.trip_dir / 'trip_info.json'
        
        if trip_info_file.exists():
            with open(trip_info_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            # 如果没有行程信息文件，返回None
            return None
    
    def generate_template(self) -> List[Tuple]:
        """
        生成报销单模板数据
        
        Returns:
            报销单数据列表 [(序号, 行程, 阶段, 出发地, 到达地, 交通方式, 金额, 发票号, 发票文件, 备注), ...]
        """
        # 尝试加载trip_info.json
        trip_info = self._load_trip_info()
        
        if trip_info:
            # 使用trip_info中的阶段信息
            outward_stages = trip_info.get('outward', [])
            return_stages = trip_info.get('return', [])
        else:
            # 如果没有trip_info，使用默认模板（兼容旧版本）
            outward_stages = [
                {"stage": 1, "from": "家", "to": "地铁站", "transport": "打车"},
                {"stage": 2, "from": "地铁站", "to": f"{self.destination}站", "transport": "地铁"},
                {"stage": 3, "from": f"{self.destination}站", "to": f"{self.destination}站", "transport": "高铁"},
                {"stage": 4, "from": f"{self.destination}站", "to": "地铁站", "transport": "地铁"},
                {"stage": 5, "from": "地铁站", "to": "宿舍", "transport": "打车"},
            ]
            return_stages = [
                {"stage": 1, "from": "宿舍", "to": "地铁站", "transport": "打车"},
                {"stage": 2, "from": "地铁站", "to": f"{self.destination}站", "transport": "地铁"},
                {"stage": 3, "from": f"{self.destination}站", "to": f"{self.departure}站", "transport": "高铁"},
                {"stage": 4, "from": f"{self.departure}站", "to": "地铁站", "transport": "地铁"},
                {"stage": 5, "from": "地铁站", "to": "家", "transport": "打车"},
            ]
        
        data = []
        seq = 1
        
        # 处理往程阶段
        for stage_info in outward_stages:
            stage = stage_info.get('stage', seq)
            from_loc = stage_info.get('from', '')
            to_loc = stage_info.get('to', '')
            transport = stage_info.get('transport', '')
            row_data = self._create_row_data(seq, '往', stage, from_loc, to_loc, transport)
            data.append(row_data)
            seq += 1
        
        # 处理返程阶段
        for stage_info in return_stages:
            stage = stage_info.get('stage', seq)
            from_loc = stage_info.get('from', '')
            to_loc = stage_info.get('to', '')
            transport = stage_info.get('transport', '')
            row_data = self._create_row_data(seq, '返', stage, from_loc, to_loc, transport)
            data.append(row_data)
            seq += 1
        
        return data
    
    def _create_row_data(self, seq: int, trip_type: str, stage: int, from_loc: str, to_loc: str, transport: str) -> Tuple:
        """
        创建一行报销单数据
        
        Args:
            seq: 序号
            trip_type: 往/返
            stage: 阶段
            from_loc: 出发地
            to_loc: 到达地
            transport: 交通工具
            
        Returns:
            (序号, 行程, 阶段, 出发地, 到达地, 交通方式, 金额, 发票号, 发票文件, 备注)
        """
        # 阶段描述
        if stage == 1:
            if '家' in from_loc:
                stage_desc = "1.家→中转"
            elif '宿' in from_loc or '宿舍' in from_loc:
                stage_desc = "1.宿→中转"
            else:
                stage_desc = f"1.{from_loc}→中转"
        elif stage == 2:
            stage_desc = "2.中转→站"
        elif stage == 3:
            stage_desc = "3.长途交通"
        elif stage == 4:
            stage_desc = "4.到→中转"
        elif stage == 5:
            if '家' in to_loc:
                stage_desc = "5.中转→家"
            elif '宿' in to_loc or '宿舍' in to_loc:
                stage_desc = "5.中转→宿"
            else:
                stage_desc = f"5.中转→{to_loc}"
        else:
            stage_desc = f"{stage}.{from_loc}→{to_loc}"
        
        # 打车阶段匹配发票
        if '打车' in transport:
            # 找尚未使用的打车发票
            for inv in self.invoices:
                if inv['invoice_type'] in ['taxi', 'other'] and inv.get('source_file', '') not in getattr(self, '_used_invoices', set()):
                    if inv.get('amount', 0) > 0:
                        # 标记为已使用
                        if not hasattr(self, '_used_invoices'):
                            self._used_invoices = set()
                        self._used_invoices.add(inv.get('source_file', ''))
                        
                        # 获取行程单备注
                        trip_remark = self._get_trip_remark(inv)
                        
                        return (seq, trip_type, stage_desc, from_loc, to_loc, transport,
                               inv.get('amount', 0), inv.get('invoice_number', ''),
                               inv.get('source_file', ''), trip_remark)
            
            # 没有找到发票
            return (seq, trip_type, stage_desc, from_loc, to_loc, transport, 0, '', '', '')
        
        # 地铁阶段
        if '地铁' in transport:
            return (seq, trip_type, stage_desc, from_loc, to_loc, transport, 0, '', '', '')
        
        # 长途交通
        if '高铁' in transport or '火车' in transport:
            # 火车票：自行报销
            for inv in self.invoices:
                if inv.get('invoice_type') == 'train':
                    if inv.get('source_file', '') not in getattr(self, '_used_invoices', set()):
                        if inv.get('amount', 0) > 0:
                            if not hasattr(self, '_used_invoices'):
                                self._used_invoices = set()
                            self._used_invoices.add(inv.get('source_file', ''))
                            
                            return (seq, trip_type, stage_desc, from_loc, to_loc, transport,
                                   inv.get('amount', 0), inv.get('invoice_number', ''),
                                   inv.get('source_file', ''), '')
            # 没找到火车票
            return (seq, trip_type, stage_desc, from_loc, to_loc, transport, 0, '', '', '')
        
        if '飞机' in transport or '航空' in transport:
            # 飞机票：公司统一报销
            return (seq, trip_type, stage_desc, from_loc, to_loc, transport, 0, '', '', '公司统一报销')
        
        # 默认
        return (seq, trip_type, stage_desc, from_loc, to_loc, transport, 0, '', '', '')
        
        return data
    
    def create_excel(self, output_path: str = None) -> str:
        """
        生成Excel文件
        
        Args:
            output_path: 输出路径
            
        Returns:
            生成的文件路径
        """
        if output_path is None:
            output_path = self.trip_dir / '报销单.xlsx'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "报销单"
        
        # 样式
        header_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        sub_header_font = Font(name='微软雅黑', size=11, bold=True)
        sub_header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        normal_font = Font(name='微软雅黑', size=10)
        border_thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        # 标题行
        ws.merge_cells('A1:J1')
        title_text = f"报销单 - {self.departure}→{self.destination} ({self.start_date} 至 {self.end_date})"
        ws['A1'] = title_text
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 30
        
        # 表头
        headers = ['序号', '行程', '阶段', '出发地', '到达地', '交通方式', '金额(元)', '发票号', '发票文件', '备注']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.value = header
            cell.font = sub_header_font
            cell.fill = sub_header_fill
            cell.alignment = center_align
            cell.border = border_thin
        
        # 生成数据
        data = self.generate_template()
        
        # 写入数据行
        for row_idx, row_data in enumerate(data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = normal_font
                cell.border = border_thin
                
                if col_idx in [1, 2, 3, 6]:  # 序号、行程、阶段、交通方式
                    cell.alignment = center_align
                elif col_idx == 7:  # 金额
                    cell.alignment = right_align
                else:
                    cell.alignment = left_align
        
        # 合计行
        row = len(data) + 3
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "合计"
        ws[f'A{row}'].font = Font(name='微软雅黑', size=11, bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws[f'G{row}'] = f"=SUM(G3:G{row-1})"
        ws[f'G{row}'].font = Font(name='微软雅黑', size=11, bold=True)
        ws[f'G{row}'].number_format = '0.00'
        ws[f'G{row}'].alignment = right_align
        
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = border_thin
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # 列宽
        widths = [6, 8, 12, 12, 12, 10, 12, 20, 25, 20]
        for col_idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # 保存
        wb.save(output_path)
        return str(output_path)


def get_column_letter(col_idx):
    """获取列字母"""
    result = ''
    while col_idx > 0:
        col_idx -= 1
        result = chr(col_idx % 26 + 65) + result
        col_idx //= 26
    return result


def generate_expense_report(trip_dir: str, output_path: str = None) -> str:
    """
    生成报销单
    
    Args:
        trip_dir: 行程目录
        output_path: 输出Excel路径
        
    Returns:
        生成的文件路径
    """
    generator = ExpenseExcelGenerator(trip_dir)
    return generator.create_excel(output_path)


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("用法: python expense_excel_generator.py <行程目录>")
        sys.exit(1)
    
    trip_dir = sys.argv[1]
    output_path = generate_expense_report(trip_dir)
    print(f"报销单已生成: {output_path}")
