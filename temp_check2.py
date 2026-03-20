# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook(r'D:\Users\luzhe\报销凭证\南京-北京-0225-0311\报销单_自动3.xlsx')
ws = wb.active

print('Excel auto-generated:')
for row in range(1, 16):
    a = ws[f'A{row}'].value or ''
    b = ws[f'B{row}'].value or ''
    c = ws[f'C{row}'].value or ''
    g = ws[f'G{row}'].value or ''
    i = ws[f'I{row}'].value or ''
    j = ws[f'J{row}'].value or ''
    print(f'Row {row}: Seq={a} Trip={b} Stage={c} Amount={g} File={i} Remark={j}')
